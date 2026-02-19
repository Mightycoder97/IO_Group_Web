"""
Diagnostic script: Inspect actual header structure & data of problematic sheets.
Outputs details about detected headers and why services might not be extractable.
"""
import openpyxl
import re
from datetime import datetime

try:
    from unidecode import unidecode as unidecode_fn
except ImportError:
    def unidecode_fn(s):
        return s

INPUT_FILE = 'PAGOS CLIENTES LIMA.xlsx'

def normalize_text(text):
    if not text: return ""
    return unidecode_fn(str(text)).upper().strip()

print(f"Cargando '{INPUT_FILE}'...")
wb = openpyxl.load_workbook(INPUT_FILE, data_only=True, read_only=True)
sheet_names = wb.sheetnames
print(f"Total hojas: {len(sheet_names)}")

# Categorize results
no_header = []
header_no_services = []
header_with_services = []

# Check ALL sheets
for sheet_name in sheet_names:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True, max_row=100))
    if not rows:
        no_header.append((sheet_name, "EMPTY"))
        continue

    # Try to find header
    header_row_idx = -1
    col_map = {}
    
    for i, row in enumerate(rows[:15]):
        row_str = " ".join([str(c) for c in row if c]).upper()
        norm_row = [normalize_text(c) for c in row]
        
        if 'FECHA' in row_str and ('SERVICIO' in row_str or 'PAGO' in row_str):
            for idx, val in enumerate(norm_row):
                if 'FECHA' in val and 'SERVICIO' in val: col_map['FECHA_SERVICIO'] = idx
                elif 'FECHA' in val and 'PAGO' in val: col_map['FECHA_PAGO'] = idx
                elif 'MES' in val: col_map['MES'] = idx
                elif 'FORMA' in val and 'PAGO' in val: col_map['FORMA_PAGO'] = idx
                elif 'DESC' in val or 'RESIDU' in val: col_map['DESCRIPCION'] = idx
                elif 'FACTURA' in val: col_map['FACTURA'] = idx
                elif 'MANIFIESTO' in val: col_map['MANIFIESTO'] = idx
            
            if 'FECHA_SERVICIO' in col_map:
                header_row_idx = i
                break
    
    if header_row_idx == -1:
        # Try alternate search - maybe header says just "FECHA" without "SERVICIO"
        for i, row in enumerate(rows[:15]):
            norm_row = [normalize_text(c) for c in row]
            has_fecha = any('FECHA' in v for v in norm_row)
            has_multiple_cols = sum(1 for v in norm_row if v.strip()) >= 3
            if has_fecha and has_multiple_cols:
                no_header.append((sheet_name, f"POSSIBLE_ALT_HEADER at row {i+1}: {[str(c)[:30] for c in row if c]}"))
                break
        else:
            no_header.append((sheet_name, f"NO_FECHA found. Top rows: {[[str(c)[:20] for c in r if c] for r in rows[:5]]}"))
        continue
    
    # Header found, now check why services fail
    start_row = header_row_idx + 1
    fecha_col = col_map.get('FECHA_SERVICIO')
    
    # Count services and diagnose
    service_count = 0
    date_fails = 0
    sample_dates = []
    
    for r_idx, row in enumerate(rows[start_row:start_row+20]):  # Check first 20 data rows
        if not row: continue
        if fecha_col is None or fecha_col >= len(row):
            date_fails += 1
            continue
        
        cell_val = row[fecha_col]
        if cell_val is None:
            continue
        
        # Try parse
        parsed = None
        if isinstance(cell_val, datetime):
            parsed = cell_val.strftime('%Y-%m-%d')
        else:
            val_str = str(cell_val).strip()
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                try:
                    dt = datetime.strptime(val_str.split(' ')[0], fmt)
                    if 2000 < dt.year < 2100:
                        parsed = dt.strftime('%Y-%m-%d')
                        break
                except:
                    pass
        
        if parsed:
            service_count += 1
            if len(sample_dates) < 3:
                sample_dates.append(parsed)
        else:
            date_fails += 1
            if len(sample_dates) < 3:
                sample_dates.append(f"FAIL:{type(cell_val).__name__}='{cell_val}'")
    
    if service_count > 0:
        header_with_services.append((sheet_name, service_count, col_map, sample_dates))
    else:
        header_no_services.append((sheet_name, header_row_idx, col_map, fecha_col, sample_dates, date_fails,
                                    f"Header row content: {[str(c)[:30] for c in rows[header_row_idx] if c]}",
                                    f"First data row: {[str(c)[:30] for c in rows[start_row] if c] if start_row < len(rows) else 'NO_DATA'}"))

# ===== REPORT =====
print(f"\n{'='*80}")
print(f"RESULTADOS DIAGNÓSTICOS")
print(f"{'='*80}")

print(f"\n✅ HOJAS CON SERVICIOS EXTRAÍBLES: {len(header_with_services)}")
for name, count, cmap, samples in header_with_services[:10]:
    print(f"  {name}: {count} servicios, cols={list(cmap.keys())}, samples={samples}")

print(f"\n⚠️  HOJAS CON HEADER PERO SIN SERVICIOS: {len(header_no_services)}")
for item in header_no_services[:20]:
    name = item[0]
    h_idx = item[1]
    cmap = item[2]
    fecha_col_idx = item[3]
    samples = item[4]
    fails = item[5]
    header_info = item[6]
    data_info = item[7]
    print(f"\n  [{name}]")
    print(f"    Header en fila: {h_idx+1}, col_map: {cmap}")
    print(f"    fecha_col: {fecha_col_idx}, date_fails: {fails}")
    print(f"    {header_info}")
    print(f"    {data_info}")
    print(f"    date_samples: {samples}")

print(f"\n❌ HOJAS SIN HEADER DETECTADO: {len(no_header)}")
for name, reason in no_header[:10]:
    print(f"  [{name}]: {reason}")

print(f"\n{'='*80}")
print(f"TOTALES: {len(header_with_services)} OK | {len(header_no_services)} problemas | {len(no_header)} sin header")
