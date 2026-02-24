#!/usr/bin/env python3
"""
Extraer Cliente, Empresa, Sede, y ContratoServicio EXCLUSIVAMENTE
desde la hoja 'Data Completa' del Excel DATA 2026.xlsx.
"""
import openpyxl
import datetime
import os

EXCEL_PATH = '/Users/sebastianretamozo/Documents/IO_Group_Web-main/DATA 2026.xlsx'
OUTPUT_PATH = '/Users/sebastianretamozo/Documents/IO_Group_Web-main/control/database/migrations/import_data_completa_2026.sql'

def escape_sql(val):
    if val is None: return 'NULL'
    if isinstance(val, bool): return '1' if val else '0'
    if isinstance(val, (int, float)): return str(val)
    if isinstance(val, datetime.datetime) or isinstance(val, datetime.date):
        return f"'{val.strftime('%Y-%m-%d')}'"
    if isinstance(val, datetime.time): return 'NULL'
    s = str(val).strip()
    if s == '' or s.lower() == 'none': return 'NULL'
    return f"'{s.replace(chr(39), chr(92)+chr(39)).replace(chr(9), '')}'"

def escape_sql_string(val):
    if val is None: return 'NULL'
    s = str(val).strip()
    if s == '' or s.lower() == 'none': return 'NULL'
    try:
        if '.' in s: s = str(int(float(s)))
    except: pass
    return f"'{s.replace(chr(39), chr(92)+chr(39)).replace(chr(9), '')}'"

def read_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    headers = [str(ws.cell(row=1, column=col).value).strip() if ws.cell(row=1, column=col).value is not None else '' for col in range(1, ws.max_column + 1)]
    return [{headers[col_idx-1]: ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers)+1)} for row_idx in range(2, ws.max_row + 1)]

def parse_data_completa(wb):
    rows = read_sheet(wb, 'Data Completa')
    
    clientes_dict = {}       # key: ruc -> values: {nombre, dni, activo}
    empresas_dict = {}       # key: ruc -> values: {razon_social, rubro, dir, dist, prov, dep, activo}
    sedes_dict = {}          # key: (ruc, distrito, direccion) -> values: {...}
    contratos_dict = {}      # key: (id_sede, fecha) -> values: {...}
    
    for idx, row in enumerate(rows):
        ruc_raw = row.get('RUC')
        if ruc_raw is None or str(ruc_raw).strip() == '' or str(ruc_raw).strip() == 'None':
            continue
            
        ruc = str(ruc_raw).strip()
        try:
            if '.' in ruc: ruc = str(int(float(ruc)))
        except: pass
        
        status_val = str(row.get('STATUS', '')).strip().lower()
        is_active = 0 if 'inactivo' in status_val else 1
        
        # 1. CLIENTE
        # Si no tiene DNI, usamos el RUC. Si no tiene nombre_cliente, usamos RAZON SOCIAL o NOMBRE COMERCIAL.
        dni = str(row.get('DNI', '')).strip()
        nombre_cliente = str(row.get('CLIENTE', '')).strip()
        if not nombre_cliente or nombre_cliente.lower() == 'none':
            nombre_cliente = str(row.get('RAZON SOCIAL', '')).strip()
        
        # Guardo el cliente o lo actualizo si la fila nueva es activa y la anterior inactiva
        if ruc not in clientes_dict or (is_active == 1 and clientes_dict[ruc]['activo'] == 0):
            clientes_dict[ruc] = {
                'nombre': nombre_cliente,
                'dni': dni if dni and dni.lower() != 'none' else None,
                'tipo_doc': "'DNI'" if dni and dni.lower() != 'none' else "'RUC'",
                'activo': is_active
            }
            
        # 2. EMPRESA
        razon_social = str(row.get('RAZON SOCIAL', '')).strip()
        if not razon_social or razon_social.lower() == 'none':
            razon_social = nombre_cliente
            
        rubro = str(row.get('RUBRO', '')).strip()
        dir_fiscal = str(row.get('DIRECCION', '')).strip()
        dist = str(row.get('DISTRITO', '')).strip().upper()
        prov = str(row.get('PROVINCIA', 'LIMA')).strip().upper()
        dep = str(row.get('DEPARTAMENTO', 'LIMA')).strip().upper()
        
        if ruc not in empresas_dict or (is_active == 1 and empresas_dict[ruc]['activo'] == 0):
            empresas_dict[ruc] = {
                'razon_social': razon_social,
                'rubro': rubro if rubro.lower() != 'none' else None,
                'direccion': dir_fiscal if dir_fiscal.lower() != 'none' else None,
                'distrito': dist if dist.lower() != 'none' else None,
                'provincia': prov if prov.lower() != 'none' else 'LIMA',
                'departamento': dep if dep.lower() != 'none' else 'LIMA',
                'activo': is_active
            }
            
        # 3. SEDE
        # Una vez identificada la empresa (RUC), una sede es un local físico único (Distrito + Dirección).
        dir_sede = str(row.get('DIRECCION', '')).strip().upper()
        dist_sede = str(row.get('DISTRITO', '')).strip().upper()
        
        sede_key = (ruc, dist_sede, dir_sede)
        if sede_key not in sedes_dict or (is_active == 1 and sedes_dict[sede_key]['activo'] == 0):
            nombre_comercial = str(row.get('NOMBRE COMERCIAL', '')).strip()
            if not nombre_comercial or nombre_comercial.lower() == 'none':
                 nombre_comercial = razon_social
            
            tel1 = str(row.get('TELEFONO PARA PROGRAMAR', '')).strip()
            tel2 = str(row.get('TELEFONO PARA COBRAR', '')).strip()
            contacto = str(row.get('CONTACTO', '')).strip()
            ubicacion = str(row.get('UBICACIÓN', '')).strip()
            
            sedes_dict[sede_key] = {
                'nombre_comercial': nombre_comercial,
                'direccion': dir_sede if dir_sede.lower() != 'none' else None,
                'distrito': dist_sede if dist_sede.lower() != 'none' else None,
                'provincia': prov if prov.lower() != 'none' else 'LIMA',
                'departamento': dep if dep.lower() != 'none' else 'LIMA',
                'contacto_nombre': contacto if contacto.lower() != 'none' else None,
                'tel1': tel1 if tel1.lower() != 'none' else None,
                'tel2': tel2 if tel2.lower() != 'none' else None,
                'coordenadas': ubicacion if ('http' in ubicacion or ',' in ubicacion) else None,
                'activo': is_active,
                'temp_sede_id': 0 # Lo llenamos después
            }
            
        # 4. CONTRATO
        peso = row.get('LIMITE DE PESO')
        peso_clean = "NULL"
        if peso is not None and str(peso).strip() != '' and str(peso).strip().lower() != 'none':
            # Remove any text, if it's 'SIN LIMITE' treat as NULL or 0
            p_str = str(peso).strip().lower()
            try:
                peso_clean = f"'{float(p_str):.2f}'"
            except ValueError:
                peso_clean = "NULL"
                
        tarifa = row.get('TARIFA')
        tarifa_clean = "'0.00'"
        if tarifa is not None and str(tarifa).strip() != '' and str(tarifa).strip().lower() != 'none':
            t_str = str(tarifa).strip().lower()
            try:
                # remove "s/" if exists
                t_str = t_str.replace('s/', '').replace('s.', '').replace(' ', '')
                tarifa_clean = f"'{float(t_str):.2f}'"
            except ValueError:
                tarifa_clean = "'0.00'"
                
        frecuencia = str(row.get('FRECUENCIA', '')).strip().lower()
        if not frecuencia or frecuencia == 'none': frecuencia = 'mensual'
        frecuencia_map = {'mensual':'mensual', 'quincenal':'quincenal', 'semanal':'semanal', 'diario':'diario', 'bimestral':'bimestral', 'trimestral':'trimestral', 'eventual':'eventual'}
        frecuencia_sql = f"'{frecuencia_map.get(frecuencia, 'mensual')}'"
        
        fecha_inicio = row.get('FECHA INICIO CONTRATO')
        fecha_fi_sql = escape_sql(fecha_inicio) if fecha_inicio else "'2020-01-01'"
        if fecha_fi_sql == 'NULL': fecha_fi_sql = "'2020-01-01'"
        
        tipo_tarifa = 'por_servicio'
        if str(tarifa).strip().lower().find('kg') != -1: tipo_tarifa = 'por_kg'
        elif str(tarifa).strip().lower().find('fijo') != -1 or str(tarifa).strip().lower().find('mensual') != -1: tipo_tarifa = 'mensual_fijo'
        
        # Use (sede_key, fecha) as unifier just in case a sede has multiple sequential contracts. 
        # But mostly it simplifies to 1 per branch.
        contrato_key = (sede_key, fecha_fi_sql)
        if contrato_key not in contratos_dict or (is_active == 1 and contratos_dict[contrato_key]['activo'] == 0):
            contratos_dict[contrato_key] = {
                'fecha_inicio': fecha_fi_sql,
                'frecuencia': frecuencia_sql,
                'peso': peso_clean,
                'tarifa': tarifa_clean,
                'tipo_tarifa': f"'{tipo_tarifa}'",
                'activo': is_active
            }

    # Asignar IDs
    c_id = 1
    ruc_to_c_id = {}
    cliente_statements = []
    
    for ruc, data in clientes_dict.items():
        ruc_to_c_id[ruc] = c_id
        doc_val = f"'{data['dni']}'" if data['dni'] else escape_sql_string(ruc)
        cliente_statements.append(f"({c_id}, {escape_sql(data['nombre'])}, {data['tipo_doc']}, {doc_val}, {data['activo']})")
        c_id += 1
        
    e_id = 1
    ruc_to_e_id = {}
    empresa_statements = []
    
    for ruc, data in empresas_dict.items():
        ruc_to_e_id[ruc] = e_id
        c_id_match = ruc_to_c_id[ruc]
        empresa_statements.append(f"({e_id}, {c_id_match}, {escape_sql(data['razon_social'])}, {escape_sql(data['rubro'])}, {escape_sql_string(ruc)}, {escape_sql(data['direccion'])}, {escape_sql(data['distrito'])}, {escape_sql(data['provincia'])}, {escape_sql(data['departamento'])}, {data['activo']})")
        e_id += 1

    s_id = 1
    sede_statements = []
    
    for sede_key, data in sedes_dict.items():
        ruc_match = sede_key[0]
        e_id_match = ruc_to_e_id[ruc_match]
        data['temp_sede_id'] = s_id
        sede_statements.append(f"({s_id}, {e_id_match}, {escape_sql(data['nombre_comercial'])}, {escape_sql(data['direccion'])}, {escape_sql(data['distrito'])}, {escape_sql(data['provincia'])}, {escape_sql(data['departamento'])}, NULL, {escape_sql(data['coordenadas'])}, {escape_sql(data['contacto_nombre'])}, {escape_sql_string(data['tel1'])}, {escape_sql_string(data['tel2'])}, NULL, {data['activo']})")
        s_id += 1

    cs_id = 1
    contrato_statements = []
    
    for contrato_key, data in contratos_dict.items():
        sede_key_match = contrato_key[0]
        s_id_match = sedes_dict[sede_key_match]['temp_sede_id']
        contrato_statements.append(f"({cs_id}, {s_id_match}, {data['fecha_inicio']}, NULL, {data['frecuencia']}, {data['peso']}, {data['tarifa']}, {data['tipo_tarifa']}, NULL, 'Mapeado desde Data Completa', {data['activo']})")
        cs_id += 1

    # Counts
    active_sedes = sum(1 for data in sedes_dict.values() if data['activo'] == 1)
    inactive_sedes = sum(1 for data in sedes_dict.values() if data['activo'] == 0)

    print(f"--- RESUMEN DE LA EXTRACCIÓN GLOBAL DESDE 'Data Completa' ---")
    print(f"Filas Totales analizadas: {len(rows)}")
    print(f"Clientes Únicos (RUC): {len(cliente_statements)}")
    print(f"Empresas Únicas (RUC): {len(empresa_statements)}")
    print(f"Sedes Únicas (RUC + Distrito + Direccion): {len(sede_statements)}")
    print(f"  -> Sedes Activas (extraídas): {active_sedes}")
    print(f"  -> Sedes Inactivas (extraídas): {inactive_sedes}")
    print(f"Contratos Generados: {len(contrato_statements)}")
    
    return cliente_statements, empresa_statements, sede_statements, contrato_statements, sedes_dict

def generate_sql(clientes, empresas, sedes, contratos):
    lines = [
        "-- Importación EXCLUSIVA desde 'Data Completa'",
        "SET SQL_MODE = 'NO_AUTO_VALUE_ON_ZERO';",
        "SET FOREIGN_KEY_CHECKS = 0;",
        "START TRANSACTION;",
        "DELETE FROM `ContratoServicio`;",
        "DELETE FROM `Sede`;",
        "DELETE FROM `Empresa`;",
        "DELETE FROM `Cliente`;",
        "ALTER TABLE `Cliente` AUTO_INCREMENT = 1;",
        "ALTER TABLE `Empresa` AUTO_INCREMENT = 1;",
        "ALTER TABLE `Sede` AUTO_INCREMENT = 1;",
        "ALTER TABLE `ContratoServicio` AUTO_INCREMENT = 1;",
        ""
    ]
    def add_batch(table, cols, data):
        lines.append(f"-- {table} ({len(data)})")
        for i in range(0, len(data), 100):
            lines.append(f"INSERT INTO `{table}` ({cols}) VALUES\n" + ",\n".join(data[i:i+100]) + ";\n")
            
    add_batch('Cliente', 'id_cliente, nombre, tipo_documento, dni, activo', clientes)
    add_batch('Empresa', 'id_empresa, id_cliente, razon_social, rubro, ruc, direccion_fiscal, distrito, provincia, departamento, activo', empresas)
    add_batch('Sede', 'id_sede, id_empresa, nombre_comercial, direccion, distrito, provincia, departamento, referencia, coordenadas_gps, contacto_nombre, contacto_telefono, contacto_telefono_2, contacto_email, activo', sedes)
    add_batch('ContratoServicio', 'id_contrato, id_sede, fecha_inicio, fecha_fin, frecuencia, peso_limite_kg, tarifa, tipo_tarifa, doc_escaneado, observaciones, activo', contratos)
    
    lines.append("SET FOREIGN_KEY_CHECKS = 1;")
    lines.append("COMMIT;")
    return "\n".join(lines)

def main():
    print("Cargando Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    clientes, empresas, sedes, contratos, _ = parse_data_completa(wb)
    
    sql = generate_sql(clientes, empresas, sedes, contratos)
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(sql)
    print(f"\nSQL guardado con éxito en: {OUTPUT_PATH}")

if __name__ == '__main__':
    main()
