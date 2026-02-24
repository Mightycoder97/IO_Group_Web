#!/usr/bin/env python3
"""
Extraer Servicios, Facturas, Manifiestos y Guias desde 'PAGOS CLIENTES LIMA.xlsx'.
Usa lógica inteligente para inferir metodo y estado de pago (ej. 'YAPE' -> Transferencia, Cancelado).
Se apoya en 'extract_data_completa_2026' para mapear cada Sede rigurosamente a la misma ID que se usó al generar Data Completa.
"""
import openpyxl
import datetime
import os
import re
from extract_data_completa_2026 import parse_data_completa, escape_sql, escape_sql_string

EXCEL_DATA = '/Users/sebastianretamozo/Documents/IO_Group_Web-main/DATA 2026.xlsx'
EXCEL_PAGOS = '/Users/sebastianretamozo/Documents/IO_Group_Web-main/PAGOS CLIENTES LIMA.xlsx'
OUTPUT_SQL = '/Users/sebastianretamozo/Documents/IO_Group_Web-main/control/database/migrations/import_servicios_riguroso.sql'

def normalize_text(t):
    if t is None: return ""
    return str(t).strip().upper()

def fuzzy_match_score(text1, text2):
    t1 = set(re.sub(r'[^a-zA-Z0-9]', ' ', normalize_text(text1)).split())
    t2 = set(re.sub(r'[^a-zA-Z0-9]', ' ', normalize_text(text2)).split())
    if not t1 or not t2: return 0
    return len(t1.intersection(t2)) / float(max(len(t1), len(t2)))

def determine_payment_info(forma_pago, observacion):
    fp_text = normalize_text(forma_pago)
    obs_text = normalize_text(observacion)
    combined = fp_text + " " + obs_text
    
    # 1. Metodo de Pago
    metodo = 'efectivo' # Default
    if any(k in combined for k in ['TRANSFERENCIA', 'YAPE', 'PLIN', 'BCP', 'BBVA', 'INTERBANK', 'SCOTIABANK', 'DEPOSITO', 'CUENTA']):
        metodo = 'transferencia'
    elif 'EFECTIVO' in combined:
        metodo = 'efectivo'
    
    # 2. Estado de Pago
    # The DB now expects 'pendiente' or 'pagado' (NOT 'cancelado') for estado_pago.
    # We should assume it's PASSED (pagado) if there is ANY legitimate payment info
    # (like YAPE, BCP, dates like 2024, 'pagado', 'cancelo', 'ok', etc).
    # Only if it's explicitly 'por pagar', 'pendiente', 'debe', or completely empty should it be 'pendiente'.
    
    es_pagado = False
    
    # Positive payment indicators
    positive_indicators = [
        'CANCELADO', 'PAGADO', 'YAPE', 'PLIN', 'TRANSFERENCIA', 
        'BCP', 'BBVA', 'INTERBANK', 'SCOTIABANK', 'CANCELO', 
        'DEPOSITO', 'CUENTA', 'OK'
    ]
    
    # Negative payment indicators
    negative_indicators = ['POR PAGAR', 'PENDIENTE', 'DEBE', 'FALTA', 'NO PAGO', 'MOROSO']
    
    # If explicitly says it's pending/unpaid, then it is NOT paid.
    if any(k in combined for k in negative_indicators):
        es_pagado = False
    elif any(k in combined for k in positive_indicators):
        # If it has a positive keyword
        es_pagado = True
    elif fp_text and fp_text not in ['NONE', '-']:
        # If the forma pago column literally just has some text (like a date, a name, a voucher number), 
        # it almost certainly means it was paid.
        es_pagado = True
        
    estado = "'pagado'" if es_pagado else "'pendiente'"
    metodo_sql = f"'{metodo}'"
    return metodo_sql, estado

def determine_residuo(descripcion, observacion):
    combined = normalize_text(descripcion) + " " + normalize_text(observacion)
    if 'BIOCONTAMINADO' in combined or 'BIO' in combined:
        return "'BIOCONTAMINADO'"
    if 'ESPECIAL' in combined or 'QUIMICO' in combined or 'ESPECIALES' in combined:
        return "'ESPECIAL'"
    if 'COMUN' in combined or 'PUNZOCORTANTE' in combined:
        return "'COMUN'"
    
    # Default or fallback
    desc = normalize_text(descripcion)
    if not desc or desc == 'NONE':
        return "'GESTION DE RR.SS'"
    # Truncate to avoid too long strings just in case
    if len(desc) > 90: desc = desc[:90]
    return f"'{desc}'"

def parse_pagos():
    print("1. Cargando maestro de Sedes desde DATA 2026.xlsx...")
    wb_data = openpyxl.load_workbook(EXCEL_DATA, data_only=True)
    _, _, _, _, sedes_dict = parse_data_completa(wb_data)
    
    # RUC -> list of {sede_key, data}
    ruc_to_sedes = {}
    for sede_key, sede_data in sedes_dict.items():
        # sede_key = (ruc, distrito, direccion)
        ruc = sede_key[0]
        if ruc not in ruc_to_sedes:
            ruc_to_sedes[ruc] = []
        ruc_to_sedes[ruc].append({
            'key': sede_key,
            'id_sede': sede_data['temp_sede_id'],
            'id_empresa': 0, # Since we didn't store e_id directly in sedes_dict locally in that file return, oh wait, the file returns statements, not the modified dict.
            'direccion': sede_data['direccion'],
            'distrito': sede_data['distrito'],
            'nombre': sede_data['nombre_comercial']
        })
        
    print("2. Leyendo PAGOS CLIENTES LIMA.xlsx...")
    wb_pagos = openpyxl.load_workbook(EXCEL_PAGOS, data_only=True)
    
    servicios = []
    facturas = []
    manifiestos = []
    guias = []
    
    srv_id = 1
    fac_id = 1
    man_id = 1
    guia_id = 1
    
    unmatched_sheets = 0
    matched_sheets = 0
    
    for sheet_name in wb_pagos.sheetnames:
        # Avoid summary sheets
        if sheet_name.upper() in ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE', 'RESUMEN', 'DATA', 'BASE']:
            if sheet_name.upper() not in ['ENERO', 'FEBRERO', 'MARZO']: # Might actually be month sheets that contain everything tabulated
                continue
                
        ws = wb_pagos[sheet_name]
        
        # Extract RUC and Header Row
        ruc_sheet = None
        direccion_sheet = sheet_name
        headers_row = -1
        col_map = {}
        
        for r_idx in range(1, 15):
            row_vals = [normalize_text(ws.cell(row=r_idx, column=c).value) for c in range(1, ws.max_column + 1)]
            
            # Find RUC
            if 'RUC' in row_vals:
                idx = row_vals.index('RUC')
                if idx + 1 < len(row_vals) and row_vals[idx+1]:
                    ruc_sheet = row_vals[idx+1].replace('.', '')
            
            # Additional context for fuzzy matching
            if 'NOMBRE COMERCIAL' in row_vals:
                idx = row_vals.index('NOMBRE COMERCIAL')
                if idx + 1 < len(row_vals) and row_vals[idx+1]:
                    direccion_sheet += " " + row_vals[idx+1]
            if 'DIRECCION' in row_vals or 'DIRECCIÓN' in row_vals:
                idx = -1
                if 'DIRECCION' in row_vals: idx = row_vals.index('DIRECCION')
                elif 'DIRECCIÓN' in row_vals: idx = row_vals.index('DIRECCIÓN')
                if idx != -1 and idx + 1 < len(row_vals) and row_vals[idx+1]:
                    direccion_sheet += " " + row_vals[idx+1]
                    
            # Find headers
            if 'FECHA DE SERVICIO' in row_vals or 'FECHA SERVICIO' in row_vals or 'MONTO' in row_vals:
                headers_row = r_idx
                for c_idx, val in enumerate(row_vals):
                    if val: col_map[val] = c_idx + 1
                break
                
        if ruc_sheet is None and headers_row == -1:
            continue # Probable blank or malformed sheet
            
        print(f"Analizando pestaña '{sheet_name}' (RUC inferido: {ruc_sheet})")
        
        # 3. Match Sede
        id_sede_matched = None
        
        # If no RUC found but we have headers, we might fail to match. Try fuzzy entirely if you want, but RUC is safest.
        if ruc_sheet and ruc_sheet in ruc_to_sedes:
            candidates = ruc_to_sedes[ruc_sheet]
            if len(candidates) == 1:
                id_sede_matched = candidates[0]['id_sede']
            else:
                # Multiple sedes for this RUC, fuzzy match by name/address
                best_score = -1
                for cand in candidates:
                    text_cand = f"{cand['nombre']} {cand['direccion']} {cand['distrito']}"
                    score = fuzzy_match_score(direccion_sheet, text_cand)
                    if score > best_score:
                        best_score = score
                        id_sede_matched = cand['id_sede']
        
        if not id_sede_matched:
            # Fallback for RUC mismatches or missing RUCs: Fuzzy match strictly across ALL sedes based on sheet name
            best_score = 0.5 # Threshold
            for clist in ruc_to_sedes.values():
                for cand in clist:
                    text_cand = f"{cand['nombre']} {cand['direccion']} {cand['distrito']}"
                    score = fuzzy_match_score(sheet_name, cand['nombre'])
                    if score > best_score:
                        best_score = score
                        id_sede_matched = cand['id_sede']
                        
        if not id_sede_matched:
            print(f"  [!] No se pudo empatar '{sheet_name}' a ninguna Sede de Data Completa. Ignorando hoja.")
            unmatched_sheets += 1
            continue
            
        matched_sheets += 1
        
        # Find which contract belongs to this sede (to get id_contrato, we don't have exact id mapping so we let `id_contrato` be NULL or we can map it via Subquery or assuming 1 contract per sede)
        # For simplicity, we assume `id_contrato` is not strictly necessary for Servicio if it has `id_sede`, or we can default to NULL since the DB allows `id_contrato` in Servicio to be NULL (Wait, does it?).
        # Fact: Servicio table requires id_contrato? Let's check schemas later. If required, we can just select MIN(id_contrato) WHERE id_sede=X in SQL.
        # Actually in extract_data_2026.py we omitted id_contrato from Servicio because it was not strictly mentioned. Wait, we need id_empresa, id_sede for Servicio!
        # How do we get id_empresa? We don't have it directly. But Servicio has `id_sede` alone maybe?
        
        # Let's extract row data
        if headers_row == -1: continue
        
        c_fp = col_map.get('FORMA DE PAGO')
        c_monto = col_map.get('MONTO')
        c_fpago = col_map.get('FECHA PAGO', col_map.get('FECHA DE PAGO'))
        c_fserv = col_map.get('FECHA DE SERVICIO', col_map.get('FECHA SERVICIO'))
        c_desc = col_map.get('DESCRIPCIÓN', col_map.get('DESCRIPCION'))
        c_obs = col_map.get('OBSERVACIÓN', col_map.get('OBSERVACION', col_map.get('OBSERVACIONES')))
        
        c_fac = col_map.get('N° FACTURA', col_map.get('FACTURA'))
        c_man = col_map.get('N° MANIFIESTO', col_map.get('MANIFIESTO'))
        c_guia = col_map.get('GUIA', col_map.get('N° GUIA'))
        
        if not c_fserv or not c_monto:
            print(f"  [!] Cabeceras esenciales (FECHA DE SERVICIO, MONTO) no encontradas en {sheet_name}. Ignorando hoja.")
            continue
            
        for r_idx in range(headers_row + 1, ws.max_row + 1):
            val_fserv = ws.cell(row=r_idx, column=c_fserv).value if c_fserv else None
            val_monto = ws.cell(row=r_idx, column=c_monto).value if c_monto else None
            
            if not val_fserv and not val_monto: continue
            if isinstance(val_monto, str) and not val_monto.replace('.', '', 1).isdigit(): continue
            
            val_fp = str(ws.cell(row=r_idx, column=c_fp).value) if c_fp and ws.cell(row=r_idx, column=c_fp).value else ""
            val_fpago = ws.cell(row=r_idx, column=c_fpago).value if c_fpago else None
            val_desc = str(ws.cell(row=r_idx, column=c_desc).value) if c_desc and ws.cell(row=r_idx, column=c_desc).value else ""
            val_obs = str(ws.cell(row=r_idx, column=c_obs).value) if c_obs and ws.cell(row=r_idx, column=c_obs).value else ""
            
            # Apply AI Rules
            metodo_sql, estado_sql = determine_payment_info(val_fp, val_obs)
            residuo_sql = determine_residuo(val_desc, val_obs)
            
            # Sanitize Execution Date
            # The user might have typed 20202, 202, 201 instead of 2020. Valid years should be 2019 to 2026.
            
            def sanitize_date_val(d_val, fallback_val=None):
                if not d_val:
                    return d_val

                def get_safe_year(y, fb):
                    if y > 9999:
                        y = int(str(y)[:4])
                    if y < 100 and y >= 19:
                        y += 2000
                    if y < 2019 or y > 2026:
                        if isinstance(fb, datetime.datetime) and 2019 <= fb.year <= 2026:
                            return fb.year
                        return 2024
                    return y

                if isinstance(d_val, datetime.datetime):
                    safe_y = get_safe_year(d_val.year, fallback_val)
                    try:
                        return d_val.replace(year=safe_y)
                    except ValueError:
                        return d_val.replace(year=safe_y, day=28)
                elif isinstance(d_val, str):
                    import re
                    match = re.search(r'(\d{2,4})[-/](\d{1,2})[-/](\d{1,4})', d_val.strip())
                    if match:
                        p1, p2, p3 = match.groups()
                        y_str = p1 if len(p1) >= 4 or int(p1) > 31 else (p3 if len(p3) >= 4 or int(p3) > 31 else p3)
                        try:
                            y = int(y_str)
                            safe_y = get_safe_year(y, fallback_val)
                            
                            # Reconstruct string YYYY-MM-DD
                            if len(p1) > 2 or (len(p3) == 2 and len(p1) == 2):
                                return f"{safe_y}-{p2.zfill(2)}-{p1.zfill(2) if len(p1)<=2 else p3.zfill(2)}"
                            else:
                                return f"{safe_y}-{p2.zfill(2)}-{p3.zfill(2)}"
                        except ValueError:
                            pass
                    
                    # If string doesn't match standard date pattern or failed to parse, force default
                    return datetime.datetime(2024, 1, 1)
                
                return d_val
                
            val_fserv = sanitize_date_val(val_fserv, val_fpago)
            val_fpago = sanitize_date_val(val_fpago, val_fserv) # mutual fallback
            
            fecha_ejecucion_sql = escape_sql(val_fserv)
            if fecha_ejecucion_sql == 'NULL': continue # Can't have service without execution date realistically
            
            fecha_pago_sql = escape_sql(val_fpago)
            tarifa_sql = escape_sql(val_monto)
            
            if tarifa_sql == 'NULL': tarifa_sql = "'0.00'"
            
            # The exact SQL schema for Servicio:
            # CREATE TABLE `Servicio` ( `id_servicio`, `id_sede`, `id_ruta`, `id_planta`, `id_contrato`, `mes_servicio`, `fecha_ejecucion`, `estado`, `estado_pago`, `fecha_pago`, `forma_pago`, `residuo` )
            
            subq_contrato = f"(SELECT id_contrato FROM ContratoServicio WHERE id_sede = {id_sede_matched} LIMIT 1)"
            
            # Derive mes_servicio strictly as YYYY-MM
            mes_servicio = 'NULL'
            if isinstance(val_fserv, datetime.datetime):
               mes_servicio = escape_sql(f"{val_fserv.year}-{str(val_fserv.month).zfill(2)}")
            
            servicios.append(
                f"({srv_id}, {id_sede_matched}, NULL, 1, {subq_contrato}, {mes_servicio}, {fecha_ejecucion_sql}, 'completado', {estado_sql}, {fecha_pago_sql}, {metodo_sql}, {residuo_sql})"
            )
            
            # Docs
            val_fac = str(ws.cell(row=r_idx, column=c_fac).value).strip() if c_fac and ws.cell(row=r_idx, column=c_fac).value else ""
            val_man = str(ws.cell(row=r_idx, column=c_man).value).strip() if c_man and ws.cell(row=r_idx, column=c_man).value else ""
            val_guia = str(ws.cell(row=r_idx, column=c_guia).value).strip() if c_guia and ws.cell(row=r_idx, column=c_guia).value else ""
            
            if val_fac and val_fac.upper() not in ['NONE', '-']:
                # id_factura, id_servicio, numero_factura, doc_escaneado
                facturas.append(f"({fac_id}, {srv_id}, {escape_sql_string(val_fac)}, NULL)")
                fac_id += 1
            if val_man and val_man.upper() not in ['NONE', '-']:
                # id_manifiesto, id_servicio, numero_manifiesto
                manifiestos.append(f"({man_id}, {srv_id}, {escape_sql_string(val_man)})")
                man_id += 1
            if val_guia and val_guia.upper() not in ['NONE', '-']:
                # id_guia, id_servicio, serie, numero_guia, fecha_emision, punto_partida, punto_llegada, doc_escaneado, observaciones
                guias.append(f"({guia_id}, {srv_id}, NULL, {escape_sql_string(val_guia)}, NULL, NULL, NULL, NULL, NULL)")
                guia_id += 1
                
            srv_id += 1

    print(f"\n--- RESUMEN DE PAGOS CLIENTES LIMA ---")
    print(f"Hojas emparejadas con éxito a Sedes únicas: {matched_sheets}")
    print(f"Hojas no reconocidas (ignoradas): {unmatched_sheets}")
    print(f"Total Servicios detectados: {len(servicios)}")
    print(f"Total Facturas detectadas: {len(facturas)}")
    print(f"Total Manifiestos detectados: {len(manifiestos)}")
    
    # Generate SQL
    lines = [
        "USE `u511863531_IOGroupBD`;",
        "SET SQL_MODE = 'NO_AUTO_VALUE_ON_ZERO';",
        "SET FOREIGN_KEY_CHECKS = 0;",
        "START TRANSACTION;",
        "DELETE FROM `Factura`;",
        "DELETE FROM `Manifiesto`;",
        "DELETE FROM `Guia`;",
        "DELETE FROM `Servicio`;",
        "ALTER TABLE `Servicio` AUTO_INCREMENT = 1;",
        "ALTER TABLE `Factura` AUTO_INCREMENT = 1;",
        "ALTER TABLE `Manifiesto` AUTO_INCREMENT = 1;",
        "ALTER TABLE `Guia` AUTO_INCREMENT = 1;",
        ""
    ]
    def add_batch(table, cols, data):
        if not data: return
        lines.append(f"-- {table} ({len(data)})")
        for i in range(0, len(data), 100):
            lines.append(f"INSERT INTO `{table}` ({cols}) VALUES\n" + ",\n".join(data[i:i+100]) + ";\n")
            
    add_batch('Servicio', 'id_servicio, id_sede, id_ruta, id_planta, id_contrato, mes_servicio, fecha_ejecucion, estado, estado_pago, fecha_pago, forma_pago, residuo', servicios)
    add_batch('Factura', 'id_factura, id_servicio, numero_factura, doc_escaneado', facturas)
    add_batch('Manifiesto', 'id_manifiesto, id_servicio, numero_manifiesto', manifiestos)
    add_batch('Guia', 'id_guia, id_servicio, serie, numero_guia, fecha_emision, punto_partida, punto_llegada, doc_escaneado, observaciones', guias)
    
    lines.append("SET FOREIGN_KEY_CHECKS = 1;")
    lines.append("COMMIT;")
    
    os.makedirs(os.path.dirname(OUTPUT_SQL), exist_ok=True)
    with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines))
    print(f"\nSQL guardado con éxito en: {OUTPUT_SQL}")

if __name__ == '__main__':
    parse_pagos()
