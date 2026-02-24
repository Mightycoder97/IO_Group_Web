import openpyxl

wb = openpyxl.load_workbook('/Users/sebastianretamozo/Documents/IO_Group_Web-main/DATA 2026.xlsx', data_only=True)
ws = wb['Data Completa']
headers = [str(ws.cell(row=1, column=col).value).strip() if ws.cell(row=1, column=col).value is not None else '' for col in range(1, ws.max_column + 1)]
rows = [{headers[col_idx-1]: ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers)+1)} for row_idx in range(2, ws.max_row + 1)]

total_rows = len(rows)

active_rows = 0
inactive_rows = 0

unique_active_sedes = set()
unique_inactive_sedes = set()
all_unique_sedes = set()

for row in rows:
    ruc_raw = row.get('RUC')
    if ruc_raw is None: continue
    ruc_str = str(ruc_raw).strip()
    try:
        if '.' in ruc_str: ruc_str = str(int(float(ruc_str)))
    except: pass
    
    distrito = str(row.get('DISTRITO', '')).strip().upper()
    if not distrito or distrito == 'NONE': continue
    
    status_val = str(row.get('STATUS', '')).strip().lower()
    
    key = (ruc_str, distrito)
    all_unique_sedes.add(key)
    
    if 'inactivo' in status_val:
        inactive_rows += 1
        unique_inactive_sedes.add(key)
    elif 'activo' in status_val:
        active_rows += 1
        unique_active_sedes.add(key)

print(f"Total Rows: {total_rows}")
print(f"Active Rows (STATUS='activo'): {active_rows}")
print(f"Inactive Rows (STATUS='inactivo'): {inactive_rows}")
print("---")
print(f"Unique Active Sedes (RUC+Distrito) in Data Completa: {len(unique_active_sedes)}")
print(f"Unique Inactive Sedes (RUC+Distrito) in Data Completa: {len(unique_inactive_sedes)}")
print(f"Total Unique Sedes in Data Completa: {len(all_unique_sedes)}")
