#!/usr/bin/env python3
"""
IO Group - Extractor Quirúrgico de Servicios 2026
Extrae los servicios de Enero a Junio de 2026 para LIMA y SUR,
emparejándolos rigurosamente solo con sedes activas en la base de datos.
"""
import openpyxl
import datetime
import re
import glob
import os
import json
from collections import defaultdict

# Setup paths relative to the script location
DB_SQL_PATH = 'control/database/migrations/import_base_datos_2026.sql'
OUTPUT_SQL_PATH = 'control/database/migrations/import_servicios_2026_lima_sur.sql'

def clean_text(val):
    if val is None: return ""
    return str(val).strip()

def clean_numeric_ruc(val):
    text = clean_text(val)
    # remove decimals like .0
    if text.endswith('.0'):
        text = text[:-2]
    return ''.join(c for c in text if c.isdigit())

def is_valid_doc_number(val):
    val_upper = clean_text(val).upper().strip()
    if not val_upper or val_upper in ['NONE', '-', 'NULL', '']:
        return False
    # Check for common text remarks in document fields
    if any(k in val_upper for k in ['CANCELADO', 'YAPE', 'PLIN', 'EFECTIVO', 'PAGADO', 'DEBE', 'MOROSO', 'PENDIENTE', 'FALTA', 'OK', 'DIFERIDA', 'ATENDIDO']):
        return False
    # Must contain at least one digit to be a document number
    if not any(c.isdigit() for c in val_upper):
        return False
    return True

def parse_date(val):
    res_date = None
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        res_date = val
    elif isinstance(val, datetime.date):
        res_date = datetime.datetime(val.year, val.month, val.day)
    else:
        text = clean_text(val)
        # common formats
        for fmt in ('%d/%m/%y', '%d/%m/%Y', '%Y-%m-%d', '%d.%m.%y', '%d.%m.%Y'):
            try:
                res_date = datetime.datetime.strptime(text, fmt)
                break
            except ValueError:
                pass
                
        if not res_date:
            # try parsing DD.MM.YY with regex
            match = re.match(r'(\d{1,2})[\./-](\d{1,2})[\./-](\d{2,4})', text)
            if match:
                p1, p2, p3 = match.groups()
                try:
                    day = int(p1)
                    month = int(p2)
                    year = int(p3)
                    if year < 100:
                        year += 2000
                    res_date = datetime.datetime(year, month, day)
                except ValueError:
                    pass
                    
    if res_date:
        # Typo correction: if year is in the future (>2026) or implausible past (<2018), correct to 2026
        if res_date.year > 2026 or res_date.year < 2018:
            try:
                res_date = res_date.replace(year=2026)
            except ValueError:
                if res_date.month == 2 and res_date.day == 29:
                    res_date = res_date.replace(year=2026, day=28)
        return res_date
    return None

def escape_sql_string(val):
    if val is None:
        return "NULL"
    cleaned = str(val).replace("'", "''").replace("\\", "\\\\")
    return f"'{cleaned}'"

def determine_payment_info(forma_pago, observacion):
    fp_text = clean_text(forma_pago).upper()
    obs_text = clean_text(observacion).upper()
    combined = fp_text + " " + obs_text
    
    # 1. Metodo de Pago
    metodo = 'efectivo' # Default
    if any(k in combined for k in ['TRANSFERENCIA', 'YAPE', 'PLIN', 'BCP', 'BBVA', 'INTERBANK', 'SCOTIABANK', 'DEPOSITO', 'CUENTA']):
        metodo = 'transferencia'
    elif 'EFECTIVO' in combined:
        metodo = 'efectivo'
    
    # 2. Estado de Pago
    es_pagado = False
    positive_indicators = ['CANCELADO', 'PAGADO', 'YAPE', 'PLIN', 'TRANSFERENCIA', 'BCP', 'BBVA', 'CANCELO', 'DEPOSITO', 'OK']
    negative_indicators = ['POR PAGAR', 'PENDIENTE', 'DEBE', 'FALTA', 'NO PAGO']
    
    if any(k in combined for k in negative_indicators):
        es_pagado = False
    elif any(k in combined for k in positive_indicators):
        es_pagado = True
    elif fp_text and fp_text not in ['NONE', '-', '']:
        es_pagado = True
        
    estado = "'pagado'" if es_pagado else "'pendiente'"
    metodo_sql = f"'{metodo}'"
    return metodo_sql, estado

def determine_residuo(descripcion, observacion):
    combined = clean_text(descripcion).upper() + " " + clean_text(observacion).upper()
    if 'BIOCONTAMINADO' in combined or 'BIO' in combined:
        return "'BIOCONTAMINADO'"
    if 'ESPECIAL' in combined or 'QUIMICO' in combined or 'ESPECIALES' in combined:
        return "'ESPECIAL'"
    if 'COMUN' in combined or 'PUNZOCORTANTE' in combined:
        return "'COMUN'"
    
    desc = clean_text(descripcion)
    if not desc or desc.upper() in ['NONE', '-']:
        return "'BIOCONTAMINADO'" # standard default for IO Group
    if len(desc) > 90: desc = desc[:90]
    return escape_sql_string(desc)

def load_sedes_from_sql(sql_path):
    print(f"Leyendo base de datos desde {sql_path}...")
    if not os.path.exists(sql_path):
        raise FileNotFoundError(f"SQL file not found at: {sql_path}")
        
    with open(sql_path) as f:
        sql = f.read()
        
    # Extract Empresa inserts: (id_empresa, id_cliente, razon_social, rubro, ruc, ...)
    empresa_to_ruc = {}
    emp_matches = re.findall(r'\((\d+),\s*(\d+),\s*\'([^\']*)\'\s*,\s*\'[^\']*\'\s*,\s*\'(\d+)\'', sql)
    for emp_id, cli_id, r_social, ruc in emp_matches:
        empresa_to_ruc[int(emp_id)] = {
            'ruc': ruc,
            'razon_social': r_social
        }
        
    # Extract Sede inserts: (id_sede, id_empresa, nombre_comercial, direccion, distrito, ...)
    # Insertion format: INSERT INTO `Sede` (`id_sede`, `id_empresa`, `nombre_comercial`, `direccion`, `distrito`, ...) VALUES
    # Let's match all rows in the Sede inserts
    sedes = []
    # Pattern matching tuples: (id_sede, id_empresa, 'nombre_comercial', 'direccion', 'distrito', ...)
    sede_tuples = re.findall(r'\((\d+),\s*(\d+),\s*\'([^\']*)\'\s*,\s*\'([^\']*)\'\s*,\s*\'([^\']*)\'', sql)
    
    for id_sede, id_empresa, n_comercial, direccion, distrito in sede_tuples:
        emp_id = int(id_empresa)
        if emp_id in empresa_to_ruc:
            emp_data = empresa_to_ruc[emp_id]
            sedes.append({
                'id_sede': int(id_sede),
                'id_empresa': emp_id,
                'ruc': emp_data['ruc'],
                'razon_social': emp_data['razon_social'],
                'nombre_comercial': n_comercial,
                'direccion': direccion,
                'distrito': distrito
            })
            
    print(f"Se cargaron {len(sedes)} sedes activas de la base de datos.")
    return sedes

def main():
    try:
        sedes = load_sedes_from_sql(DB_SQL_PATH)
    except Exception as e:
        print(f"Error al cargar sedes: {e}")
        return

    # Index sedes by RUC for fast matching.
    # Note: A single RUC can have multiple sedes (venues).
    sedes_by_ruc = defaultdict(list)
    for s in sedes:
        sedes_by_ruc[s['ruc']].append(s)

    # Output structures
    servicios = []
    manifiestos = []
    guias = []
    facturas = []
    
    # Safe offsets to avoid colliding with existing database service/doc records
    # Max existing IDs: Servicio = 44078, Manifiesto = 71390, Guia = 2180, Factura = 71390
    srv_id = 50000
    man_id = 80000
    guia_id = 3000
    fac_id = 80000

    files_to_scan = [
        ('PAGOS CLIENTES LIMA.xlsx', 'LIMA'),
        ('PAGOS CLIENTES SUR.xlsx', 'SUR')
    ]
    
    skipped_sheets_no_ruc = 0
    skipped_sheets_inactive = 0
    processed_sheets = 0
    total_services_extracted = 0

    for filepath, region in files_to_scan:
        if not os.path.exists(filepath):
            print(f"Advertencia: Archivo {filepath} no encontrado. Saltando...")
            continue
            
        print(f"\nProcesando {filepath} ({region})...")
        wb = openpyxl.load_workbook(filepath, read_only=True)
        
        for sname in wb.sheetnames:
            sheet = wb[sname]
            row_iter = sheet.iter_rows(values_only=True)
            
            # Read first few rows to extract RUC and client name
            ruc_val = None
            client_title = None
            try:
                row0 = next(row_iter)
                client_title = clean_text(row0[0]) if row0 and len(row0) > 0 else None
                for _ in range(5):
                    row = next(row_iter)
                    if not row: continue
                    for idx, val in enumerate(row):
                        if val is not None and 'RUC' in str(val).upper():
                            if idx + 1 < len(row) and row[idx + 1] is not None:
                                ruc_val = str(row[idx + 1]).strip()
            except StopIteration:
                pass
                
            ruc_clean = clean_numeric_ruc(ruc_val)
            if not ruc_clean:
                skipped_sheets_no_ruc += 1
                continue
                
            # Quirurgical match: Only import if the RUC exists in our active DB sedes
            matched_sedes = sedes_by_ruc.get(ruc_clean)
            if not matched_sedes:
                skipped_sheets_inactive += 1
                continue
                
            # If multiple sedes exist for the same RUC, resolve which one matches by district/address
            selected_sede = matched_sedes[0]
            if len(matched_sedes) > 1:
                # Try to match sheet name or meta information to the Sede's district or address
                best_score = -1
                for ms in matched_sedes:
                    score = 0
                    dist = ms['distrito'].upper()
                    addr = ms['direccion'].upper()
                    sh_upper = sname.upper()
                    
                    if dist and dist in sh_upper:
                        score += 5
                    if addr and any(word in sh_upper for word in addr.split() if len(word) > 3):
                        score += 3
                        
                    if score > best_score:
                        best_score = score
                        selected_sede = ms

            # Reset row iterator to parse services starting from row 6
            row_iter = sheet.iter_rows(values_only=True)
            for _ in range(6):
                try: next(row_iter)
                except StopIteration: break
                
            processed_sheets += 1
            sheet_services_count = 0
            
            for row in row_iter:
                if not row or len(row) < 4:
                    continue
                
                val_fpago = row[0]
                val_monto = row[1]
                val_mes = row[2]
                val_fserv = row[3]
                val_desc = row[4] if len(row) > 4 else "GESTION DE RR.SS"
                val_forma = row[5] if len(row) > 5 else None
                val_nserv = row[6] if len(row) > 6 else None
                val_fac = row[7] if len(row) > 7 else None
                val_contr = row[8] if len(row) > 8 else None
                val_man = row[9] if len(row) > 9 else None
                # Guide number is sometimes at index 10 or in observations.
                val_guia = row[10] if len(row) > 10 else None
                val_obs = row[11] if len(row) > 11 else ""
                
                dt_fserv = parse_date(val_fserv)
                if not dt_fserv:
                    continue # can't parse service date
                    
                # Surgical filter: Only process services in the first half of 2026 (Jan - Jun)
                if dt_fserv.year != 2026 or dt_fserv.month > 6:
                    continue
                    
                dt_fpago = parse_date(val_fpago)
                if dt_fpago and dt_fpago.year != 2026:
                    try:
                        dt_fpago = dt_fpago.replace(year=2026)
                    except ValueError:
                        if dt_fpago.month == 2 and dt_fpago.day == 29:
                            dt_fpago = dt_fpago.replace(year=2026, day=28)
                
                # Format dates and text
                fecha_ejecucion_sql = f"'{dt_fserv.strftime('%Y-%m-%d')}'"
                fecha_pago_sql = f"'{dt_fpago.strftime('%Y-%m-%d')}'" if dt_fpago else "NULL"
                
                # Derive mes_servicio YYYY-MM
                mes_servicio_sql = f"'{dt_fserv.year}-{str(dt_fserv.month).zfill(2)}'"
                
                # Tariff/Monto
                try:
                    monto_val = float(str(val_monto).replace('S/', '').replace(',', '').strip()) if val_monto is not None else 0.0
                    monto_sql = f"{monto_val:.2f}"
                except ValueError:
                    monto_sql = "NULL"
                    
                # Payment method and status
                metodo_sql, estado_sql = determine_payment_info(val_forma, val_obs)
                
                # Residuo type/description
                residuo_sql = determine_residuo(val_desc, val_obs)
                
                # Observations
                obs_sql = escape_sql_string(val_obs) if val_obs else "NULL"
                
                # Sede Match Details
                id_sede_matched = selected_sede['id_sede']
                subq_contrato = f"(SELECT id_contrato FROM ContratoServicio WHERE id_sede = {id_sede_matched} AND activo = 1 LIMIT 1)"
                
                # Append SQL row
                servicios.append(
                    f"({srv_id}, {id_sede_matched}, NULL, 1, {subq_contrato}, {mes_servicio_sql}, {fecha_ejecucion_sql}, 'completado', {estado_sql}, {fecha_pago_sql}, {metodo_sql}, {residuo_sql}, {obs_sql}, {monto_sql})"
                )
                
                # Handle documents linked to this service
                # 1. Factura
                fact_num = clean_text(val_fac)
                if fact_num and is_valid_doc_number(fact_num):
                    facturas.append(f"({fac_id}, {srv_id}, {escape_sql_string(fact_num)}, NULL)")
                    fac_id += 1
                    
                # 2. Manifiesto
                man_num = clean_text(val_man)
                if man_num and is_valid_doc_number(man_num):
                    manifiestos.append(f"({man_id}, {srv_id}, {escape_sql_string(man_num)})")
                    man_id += 1
                    
                # 3. Guia de Transporte
                guia_num = clean_text(val_guia)
                if guia_num and is_valid_doc_number(guia_num):
                    guias.append(f"({guia_id}, {srv_id}, NULL, {escape_sql_string(guia_num)}, NULL, NULL, NULL, NULL, NULL)")
                    guia_id += 1
                
                srv_id += 1
                sheet_services_count += 1
                total_services_extracted += 1
            
            if sheet_services_count > 0:
                # Debug print
                pass
                
    # 5. Write SQL Migration File
    print(f"\nGenerando script SQL en {OUTPUT_SQL_PATH}...")
    
    sql_lines = []
    sql_lines.append("-- SQL Migration for IO Group - Surgical Services 2026 (Lima & Sur)")
    sql_lines.append("-- Generated on " + datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    sql_lines.append("START TRANSACTION;\n")
    
    # Enable session variable for auto increment values
    sql_lines.append("SET FOREIGN_KEY_CHECKS = 0;\n")
    
    # 1. Insert Servicios
    if servicios:
        sql_lines.append("INSERT INTO `Servicio` (`id_servicio`, `id_sede`, `id_ruta`, `id_planta`, `id_contrato`, `mes_servicio`, `fecha_ejecucion`, `estado`, `estado_pago`, `fecha_pago`, `forma_pago`, `residuo`, `observaciones`, `monto_cobrado`) VALUES")
        for i, srv_row in enumerate(servicios):
            suffix = ";" if i == len(servicios) - 1 else ","
            sql_lines.append(f"  {srv_row}{suffix}")
        sql_lines.append("")
        
    # 2. Insert Facturas
    if facturas:
        sql_lines.append("INSERT INTO `Factura` (`id_factura`, `id_servicio`, `numero_factura`, `doc_escaneado`) VALUES")
        for i, fac_row in enumerate(facturas):
            suffix = ";" if i == len(facturas) - 1 else ","
            sql_lines.append(f"  {fac_row}{suffix}")
        sql_lines.append("")

    # 3. Insert Manifiestos
    if manifiestos:
        sql_lines.append("INSERT INTO `Manifiesto` (`id_manifiesto`, `id_servicio`, `numero_manifiesto`) VALUES")
        for i, man_row in enumerate(manifiestos):
            suffix = ";" if i == len(manifiestos) - 1 else ","
            sql_lines.append(f"  {man_row}{suffix}")
        sql_lines.append("")
        
    # 4. Insert Guias
    if guias:
        sql_lines.append("INSERT INTO `Guia` (`id_guia`, `id_servicio`, `serie`, `numero_guia`, `fecha_emision`, `punto_partida`, `punto_llegada`, `doc_escaneado`, `observaciones`) VALUES")
        for i, guia_row in enumerate(guias):
            suffix = ";" if i == len(guias) - 1 else ","
            sql_lines.append(f"  {guia_row}{suffix}")
        sql_lines.append("")

    sql_lines.append("SET FOREIGN_KEY_CHECKS = 1;")
    sql_lines.append("COMMIT;")
    
    with open(OUTPUT_SQL_PATH, 'w') as out_f:
        out_f.write('\n'.join(sql_lines))
        
    print("\n=== RESUMEN DE LA EXTRACCIÓN ===")
    print(f"Hojas procesadas (Sedes Activas matched): {processed_sheets}")
    print(f"Hojas omitidas (RUC no registrado en BD activa): {skipped_sheets_inactive}")
    print(f"Hojas omitidas (Sin RUC detectado): {skipped_sheets_no_ruc}")
    print(f"Total servicios 2026 extraídos (Ene-Jun): {total_services_extracted}")
    print(f"Total facturas vinculadas: {len(facturas)}")
    print(f"Total manifiestos vinculados: {len(manifiestos)}")
    print(f"Total guías vinculadas: {len(guias)}")
    print(f"Migración escrita exitosamente en {OUTPUT_SQL_PATH}")

if __name__ == "__main__":
    main()
