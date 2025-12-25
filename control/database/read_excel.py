import openpyxl

# Cargar Excel
wb = openpyxl.load_workbook(r'c:\Users\sebas\Documents\IOGROUPWEB\PAGOS CLIENTES LIMA.xlsx', read_only=True)
ws = wb.active

# Obtener headers
headers = []
for cell in next(ws.iter_rows(min_row=1, max_row=1)):
    headers.append(cell.value)

print("=" * 60)
print("COLUMNAS DEL EXCEL:")
print("=" * 60)
for i, h in enumerate(headers):
    print(f"{i+1}. {h}")

print("\n" + "=" * 60)
print("BUSCANDO COLUMNA DE GUIA:")
print("=" * 60)
for i, h in enumerate(headers):
    if h and 'guia' in str(h).lower():
        print(f"ENCONTRADA: Columna {i+1} = '{h}'")

# Mostrar primeras 3 filas de datos
print("\n" + "=" * 60)
print("PRIMERAS 3 FILAS DE DATOS:")
print("=" * 60)
row_count = 0
for row in ws.iter_rows(min_row=2, max_row=4):
    row_count += 1
    print(f"\n--- Fila {row_count} ---")
    for i, cell in enumerate(row):
        if cell.value:
            print(f"  {headers[i] if i < len(headers) else f'Col{i}'}: {cell.value}")

wb.close()
