import openpyxl

file_path = '/Users/sebastianretamozo/Documents/IO_Group_Web-main/PAGOS CLIENTES LIMA.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
print("Sheet Names:", wb.sheetnames)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(str(ws.cell(row=1, column=col).value).strip() if ws.cell(row=1, column=col).value is not None else '')
    print(f"\nHeaders for '{sheet}':")
    print(headers)
    break  # Just check the first sheet to get the structure
