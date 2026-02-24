#!/usr/bin/env python3
"""
Extract Cliente, Empresa, Sede, and ContratoServicio data from DATA 2026.xlsx
and auto-discover missing Sedes from 'Data Completa'.
"""
import openpyxl
import datetime
import os

EXCEL_PATH = '/Users/sebastianretamozo/Documents/IO_Group_Web-main/DATA 2026.xlsx'
OUTPUT_PATH = '/Users/sebastianretamozo/Documents/IO_Group_Web-main/control/database/migrations/import_data_2026_new.sql'

def escape_sql(val):
    if val is None: return 'NULL'
    if isinstance(val, bool): return '1' if val else '0'
    if isinstance(val, (int, float)): return str(val)
    if isinstance(val, datetime.datetime) or isinstance(val, datetime.date):
        return f"'{val.strftime('%Y-%m-%d')}'"
    if isinstance(val, datetime.time): return 'NULL'
    s = str(val).strip()
    if s == '' or s.lower() == 'none': return 'NULL'
    return f"'{s.replace(chr(39), chr(92)+chr(39)).replace(chr(9), '')}'"

def escape_sql_string(val):
    if val is None: return 'NULL'
    s = str(val).strip()
    if s == '' or s.lower() == 'none': return 'NULL'
    try:
        if '.' in s: s = str(int(float(s)))
    except: pass
    return f"'{s.replace(chr(39), chr(92)+chr(39)).replace(chr(9), '')}'"

def map_frecuencia(val):
    s = str(val).strip().lower() if val else ''
    return f"'{ { 'mensual':'mensual', 'quincenal':'quincenal', 'semanal':'semanal', 'diario':'diario', 'bimestral':'bimestral', 'trimestral':'trimestral', 'eventual':'eventual' }.get(s, 'mensual') }'"

def map_tipo_tarifa(val):
    s = str(val).strip().lower() if val else ''
    if 'kg' in s: return "'por_kg'"
    if 'fijo' in s or 'mensual' in s: return "'mensual_fijo'"
    return "'por_servicio'"

def read_sheet(wb, sheet_name):
    ws = wb[sheet_name]
    headers = [str(ws.cell(row=1, column=col).value).strip() if ws.cell(row=1, column=col).value is not None else '' for col in range(1, ws.max_column + 1)]
    return [{headers[col_idx-1]: ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers)+1)} for row_idx in range(2, ws.max_row + 1)]

def extract_clientes(wb):
    rows = read_sheet(wb, 'CLIENTES')
    statements = []
    for row in rows:
        if row.get('id_cliente') is None: continue
        statements.append(f"({int(row['id_cliente'])}, {escape_sql(row.get('nombre'))}, {escape_sql(row.get('tipo_documento'))}, {escape_sql_string(row.get('dni'))}, {int(row.get('activo',1)) if row.get('activo') is not None else 1})")
    return statements, len(statements)

def extract_empresas(wb):
    rows = read_sheet(wb, 'EMPRESA')
    statements = []
    empresa_map = {}
    for row in rows:
        if row.get('id_empresa') is None: continue
        id_e = int(row['id_empresa'])
        ruc_raw = str(row.get('ruc', '')).strip()
        if '.' in ruc_raw:
            try: ruc_raw = str(int(float(ruc_raw)))
            except: pass
        if ruc_raw and ruc_raw.lower() != 'none':
            empresa_map[ruc_raw] = id_e
            
        statements.append(f"({id_e}, {int(row['id_cliente'])}, {escape_sql(row.get('razon_social'))}, NULL, {escape_sql_string(row.get('ruc'))}, {escape_sql(row.get('direccion_fiscal'))}, {escape_sql(row.get('distrito'))}, {escape_sql(row.get('provincia'))}, {escape_sql(row.get('departamento'))}, {int(row.get('activo',1)) if row.get('activo') is not None else 1})")
    return statements, len(statements), empresa_map

def parse_status_from_data_completa(wb):
    """Scan Data Completa to pick up true activo/inactivo status for every Sede."""
    rows = read_sheet(wb, 'Data Completa')
    status_map = {}
    for row in rows:
        ruc_raw = row.get('RUC')
        if ruc_raw is None: continue
        ruc_str = str(ruc_raw).strip()
        try:
            if '.' in ruc_str: ruc_str = str(int(float(ruc_str)))
        except: pass
        
        distrito = str(row.get('DISTRITO', '')).strip().upper()
        if not distrito or distrito == 'NONE': continue
        
        status_val = str(row.get('STATUS', '')).strip().lower()
        
        # Only explicitly map if there's a clear status
        if 'inactivo' in status_val:
            status_map[(ruc_str, distrito)] = 0
        elif 'activo' in status_val:
            status_map[(ruc_str, distrito)] = 1
        
    return status_map

def extract_sedes(wb, empresa_map, status_map):
    rows = read_sheet(wb, 'SEDE')
    statements = []
    sede_activo_map = {}
    sede_lookup = {}
    max_id_sede = 0
    id_empresa_to_ruc = {v: k for k, v in empresa_map.items()}
    
    for row in rows:
        if row.get('id_sede') is None or row.get('id empresa') is None: continue
        id_s = int(row['id_sede'])
        id_e = int(row['id empresa'])
        max_id_sede = max(max_id_sede, id_s)
        
        ruc = id_empresa_to_ruc.get(id_e)
        distrito = str(row.get('distrito', '')).strip().upper()
        
        # Determine activo status from Data Completa if explicit, otherwise fallback to SEDE sheet
        fallback_activo = int(row.get('activo', 1)) if row.get('activo') is not None else 1
        activo = status_map.get((ruc, distrito), fallback_activo) if ruc else fallback_activo
        
        sede_activo_map[id_s] = activo
        
        if ruc and distrito:
            sede_lookup[(ruc, distrito)] = id_s
            
        statements.append(f"({id_s}, {id_e}, {escape_sql(row.get('nombre_comercial'))}, {escape_sql(row.get('direccion'))}, {escape_sql(row.get('distrito'))}, {escape_sql(row.get('provincia'))}, {escape_sql(row.get('departamento'))}, {escape_sql(row.get('referencia'))}, {escape_sql(row.get('coordenadas_gps'))}, {escape_sql(row.get('contacto_nombre'))}, {escape_sql_string(row.get('contacto_telefono'))}, {escape_sql_string(row.get('contacto_telefono_2'))}, {escape_sql(row.get('contacto_email'))}, {activo})")
    
    return statements, len(statements), sede_activo_map, sede_lookup, max_id_sede

def extract_contratos(wb, sede_activo_map):
    rows = read_sheet(wb, 'ContratoServicio')
    statements = []
    max_id_contrato = 0
    for row in rows:
        if row.get('id_contrato') is None or row.get('id_sede') is None: continue
        id_c = int(row['id_contrato'])
        max_id_contrato = max(max_id_contrato, id_c)
        fi = escape_sql(row.get('fecha_inicio'))
        if fi == 'NULL': fi = "'2020-01-01'"
        
        peso = row.get('peso_limite_kg')
        peso_str = f"{float(peso):.2f}" if peso is not None else 'NULL'
        tarifa = row.get('tarifa')
        tarifa_str = f"{float(tarifa):.2f}" if tarifa is not None else '0.00'
        
        statements.append(f"({id_c}, {int(row['id_sede'])}, {fi}, {escape_sql(row.get('fecha_fin'))}, {map_frecuencia(row.get('frecuencia'))}, {peso_str}, {tarifa_str}, {map_tipo_tarifa(row.get('tipo_tarifa'))}, {escape_sql(row.get('doc_escaneado'))}, {escape_sql(row.get('observaciones'))}, {sede_activo_map.get(int(row['id_sede']), 1)})")
    return statements, len(statements), max_id_contrato

def discover_missing_sedes(wb, empresa_map, sede_lookup, max_id_sede, max_id_contrato, status_map):
    rows = read_sheet(wb, 'Data Completa')
    new_sedes = []
    new_contratos = []
    added = 0
    
    for row in rows:
        ruc_raw = row.get('RUC')
        if ruc_raw is None: continue
        ruc_str = str(ruc_raw).strip()
        try:
            if '.' in ruc_str: ruc_str = str(int(float(ruc_str)))
        except: pass
        
        distrito = str(row.get('DISTRITO', '')).strip().upper()
        if not distrito or distrito == 'NONE': continue
        
        key = (ruc_str, distrito)
        if key not in sede_lookup and ruc_str in empresa_map:
            id_empresa = empresa_map[ruc_str]
            max_id_sede += 1
            max_id_contrato += 1
            sede_lookup[key] = max_id_sede
            
            # For auto-discovered historical sedes: default to INACTIVE (0), unless explicitly ACTIVO (1)
            activo = status_map.get(key, 0)
            
            nombre_comercial = escape_sql(row.get('NOMBRE COMERCIAL', row.get('CLIENTE', 'Nueva Sede')))
            dir_sql = escape_sql(row.get('DIRECCION', ''))
            dist_sql = escape_sql(distrito)
            
            new_sedes.append(f"({max_id_sede}, {id_empresa}, {nombre_comercial}, {dir_sql}, {dist_sql}, 'LIMA', 'LIMA', NULL, NULL, NULL, NULL, NULL, NULL, {activo})")
            new_contratos.append(f"({max_id_contrato}, {max_id_sede}, '2020-01-01', NULL, 'mensual', NULL, 0.00, 'por_servicio', NULL, 'Generado automáticamente por script', {activo})")
            added += 1
            
    return new_sedes, new_contratos, added

def generate_sql(clientes, empresas, sedes, contratos):
    lines = [
        "-- Importación de datos",
        "USE `u511863531_IOGroupBD`;",
        "SET SQL_MODE = 'NO_AUTO_VALUE_ON_ZERO';",
        "SET FOREIGN_KEY_CHECKS = 0;",
        "START TRANSACTION;",
        "DELETE FROM `ContratoServicio`;",
        "DELETE FROM `Sede`;",
        "DELETE FROM `Empresa`;",
        "DELETE FROM `Cliente`;",
        "ALTER TABLE `Cliente` AUTO_INCREMENT = 1;",
        "ALTER TABLE `Empresa` AUTO_INCREMENT = 1;",
        "ALTER TABLE `Sede` AUTO_INCREMENT = 1;",
        "ALTER TABLE `ContratoServicio` AUTO_INCREMENT = 1;",
        ""
    ]
    def add_batch(table, cols, data):
        lines.append(f"-- {table} ({len(data)})")
        for i in range(0, len(data), 100):
            lines.append(f"INSERT INTO `{table}` ({cols}) VALUES\n" + ",\n".join(data[i:i+100]) + ";\n")
            
    add_batch('Cliente', 'id_cliente, nombre, tipo_documento, dni, activo', clientes)
    add_batch('Empresa', 'id_empresa, id_cliente, razon_social, rubro, ruc, direccion_fiscal, distrito, provincia, departamento, activo', empresas)
    add_batch('Sede', 'id_sede, id_empresa, nombre_comercial, direccion, distrito, provincia, departamento, referencia, coordenadas_gps, contacto_nombre, contacto_telefono, contacto_telefono_2, contacto_email, activo', sedes)
    add_batch('ContratoServicio', 'id_contrato, id_sede, fecha_inicio, fecha_fin, frecuencia, peso_limite_kg, tarifa, tipo_tarifa, doc_escaneado, observaciones, activo', contratos)
    
    lines.append("SET FOREIGN_KEY_CHECKS = 1;")
    lines.append("COMMIT;")
    return "\n".join(lines)

def main():
    print("Cargando Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    clientes, _ = extract_clientes(wb)
    empresas, _, empresa_map = extract_empresas(wb)
    
    print("Analizando STATUS global de Sedes...")
    status_map = parse_status_from_data_completa(wb)
    
    sedes, _, sede_activo_map, sede_lookup, max_id_sede = extract_sedes(wb, empresa_map, status_map)
    contratos, _, max_id_contrato = extract_contratos(wb, sede_activo_map)
    
    print("Buscando sedes faltantes en Data Completa...")
    n_sedes, n_contratos, added = discover_missing_sedes(wb, empresa_map, sede_lookup, max_id_sede, max_id_contrato, status_map)
    sedes.extend(n_sedes)
    contratos.extend(n_contratos)
    print(f"  -> Se auto-descubrieron y agregaron {added} sedes y contratos nuevos!")
    
    sql = generate_sql(clientes, empresas, sedes, contratos)
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(sql)
    print("SQL guardado.")

if __name__ == '__main__':
    main()

