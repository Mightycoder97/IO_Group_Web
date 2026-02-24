import openpyxl

wb = openpyxl.load_workbook('/Users/sebastianretamozo/Documents/IO_Group_Web-main/DATA 2026.xlsx', data_only=True)
ws = wb['Data Completa']
headers = [str(ws.cell(row=1, column=col).value).strip() if ws.cell(row=1, column=col).value is not None else '' for col in range(1, ws.max_column + 1)]
rows = [{headers[col_idx-1]: ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers)+1)} for row_idx in range(2, ws.max_row + 1)]

unique_active = set()
unique_inactive = set()
for r in rows:
    ruc = str(r.get('RUC', '')).strip()
    distrito = str(r.get('DISTRITO', '')).strip().upper()
    dir = str(r.get('DIRECCION', '')).strip().upper()
    nombre = str(r.get('NOMBRE COMERCIAL', '')).strip().upper()
    status = str(r.get('STATUS', '')).strip().lower()
    
    key = (ruc, distrito, dir, nombre)
    if 'inactivo' in status:
        unique_inactive.add(key)
    elif 'activo' in status:
        unique_active.add(key)
        
print(f'Active: {len(unique_active)} Inactive: {len(unique_inactive)}')
