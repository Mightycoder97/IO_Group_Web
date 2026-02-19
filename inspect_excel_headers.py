
import openpyxl

INPUT_FILE = 'PAGOS CLIENTES LIMA.xlsx'

try:
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True)
    print(f"Sheets: {wb.sheetnames[:5]}...")
    
    for sheet_name in wb.sheetnames[:3]:
        print(f"\n--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]
        
        # Method 1: Iterating rows
        rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
        
        for i, row in enumerate(rows):
            print(f"Row {i+1}:", row)
            
except Exception as e:
    print(f"Error: {e}")
