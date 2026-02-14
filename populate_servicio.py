#!/usr/bin/env python3
"""
Script to populate the Servicio table in DATA 2026.xlsx
by reading service data from PAGOS CLIENTES LIMA.xlsx,
matching empresas by RUC.
"""
import openpyxl
from datetime import datetime
import os
import shutil

BASE_DIR = '/Users/sebastianretamozo/Documents/IO_Group_Web-main'
DATA_FILE = os.path.join(BASE_DIR, 'DATA 2026.xlsx')
PAGOS_FILE = os.path.join(BASE_DIR, 'PAGOS CLIENTES LIMA.xlsx')

# --- Step 1: Load lookup tables from DATA 2026 ---
print("Loading DATA 2026.xlsx...")
wb_data_ro = openpyxl.load_workbook(DATA_FILE, read_only=True, data_only=True)

# EMPRESA: id_empresa, ruc
ws_emp = wb_data_ro['EMPRESA']
empresa_ruc = {}       # id_empresa -> ruc_str
ruc_to_empresa = {}    # ruc_str -> id_empresa
for row in list(ws_emp.iter_rows(min_row=2, values_only=True)):
    id_emp, _, _, ruc = row[0], row[1], row[2], row[3]
    if id_emp and ruc:
        ruc_str = str(int(ruc)) if isinstance(ruc, (int, float)) else str(ruc).strip()
        empresa_ruc[id_emp] = ruc_str
        ruc_to_empresa[ruc_str] = id_emp

# SEDE: id_empresa -> [(id_sede, nombre)]
ws_sede = wb_data_ro['SEDE']
empresa_sedes = {}     # id_empresa -> [(id_sede, nombre)]
sede_empresa = {}      # id_sede -> id_empresa
for row in list(ws_sede.iter_rows(min_row=2, values_only=True)):
    id_sede, id_emp, nombre = row[0], row[1], row[2]
    if id_emp and id_sede:
        if id_emp not in empresa_sedes:
            empresa_sedes[id_emp] = []
        empresa_sedes[id_emp].append((id_sede, str(nombre).strip() if nombre else ''))
        sede_empresa[id_sede] = id_emp

# ContratoServicio: id_sede -> id_contrato
ws_cont = wb_data_ro['ContratoServicio']
sede_contrato = {}     # id_sede -> id_contrato
for row in list(ws_cont.iter_rows(min_row=2, values_only=True)):
    id_contrato, id_sede = row[0], row[1]
    if id_sede and id_contrato:
        sede_contrato[id_sede] = id_contrato

wb_data_ro.close()

print(f"  Empresas: {len(empresa_ruc)}")
print(f"  Sedes: {len(sede_empresa)}")
print(f"  Contratos: {len(sede_contrato)}")

# --- Step 2: Build RUC -> sheet name index from PAGOS CLIENTES LIMA ---
print("\nLoading PAGOS CLIENTES LIMA.xlsx...")
wb_pagos = openpyxl.load_workbook(PAGOS_FILE, read_only=True, data_only=True)

ruc_to_sheet = {}  # ruc_str -> sheet_name
for sn in wb_pagos.sheetnames:
    ws = wb_pagos[sn]
    rows = list(ws.iter_rows(max_row=6, values_only=True))
    for r in rows:
        if r and len(r) > 1 and r[0]:
            cell_label = str(r[0]).strip().upper()
            if cell_label == 'RUC' and r[1]:
                ruc_val = str(int(r[1])) if isinstance(r[1], (int, float)) else str(r[1]).strip()
                ruc_to_sheet[ruc_val] = sn
                break

print(f"  PAGOS sheets with RUC: {len(ruc_to_sheet)}")

# --- Step 3: For each empresa, find the matching PAGOS sheet and extract services ---
def find_header_row(ws, max_search=15):
    """Find the row index that contains the headers (FECHA PAGO, MES DEL SERVICIO, etc.)"""
    rows = list(ws.iter_rows(max_row=max_search, values_only=True))
    for i, row in enumerate(rows):
        if row and any(cell and 'MES DEL SERVICIO' in str(cell).upper() for cell in row if cell):
            return i, row
    return None, None

def extract_services(ws, header_idx):
    """Extract service rows starting after the header row."""
    all_rows = list(ws.iter_rows(values_only=True))
    services = []
    
    # Get header to find column indices
    header = all_rows[header_idx]
    header_upper = [str(h).strip().upper() if h else '' for h in header]
    
    # Find column indices by header name
    col_map = {}
    for idx, h in enumerate(header_upper):
        if 'FECHA PAGO' in h or h == 'FECHA PAGO':
            col_map['fecha_pago'] = idx
        elif 'MES DEL SERVICIO' in h:
            col_map['mes_servicio'] = idx
        elif 'FECHA DE SERV' in h:
            col_map['fecha_ejecucion'] = idx
        elif h in ('DESCRIPCIÓN', 'DESCRIPCION'):
            col_map['descripcion'] = idx
        elif 'FORMA DE PAGO' in h:
            col_map['forma_pago'] = idx
        elif h in ('OBSERVACIÓN', 'OBSERVACION'):
            col_map['estado'] = idx
    
    # Extract data rows (after header)
    for row in all_rows[header_idx + 1:]:
        # A valid service row must have at least MES DEL SERVICIO
        mes_idx = col_map.get('mes_servicio')
        if mes_idx is None or mes_idx >= len(row) or not row[mes_idx]:
            continue
        
        mes = row[mes_idx]
        if not mes or (isinstance(mes, str) and not mes.strip()):
            continue
        
        # Also check that there's a fecha_ejecucion
        fecha_ej_idx = col_map.get('fecha_ejecucion')
        fecha_ej = row[fecha_ej_idx] if fecha_ej_idx is not None and fecha_ej_idx < len(row) else None
        if not fecha_ej:
            continue
        
        service = {
            'mes_servicio': str(mes).strip() if mes else None,
            'fecha_ejecucion': fecha_ej if isinstance(fecha_ej, datetime) else None,
            'estado': None,
            'estado_pago': None,
            'fecha_pago': None,
            'forma_pago': None,
            'descripcion_residuo': None,
        }
        
        # estado from OBSERVACION
        estado_idx = col_map.get('estado')
        if estado_idx is not None and estado_idx < len(row) and row[estado_idx]:
            service['estado'] = str(row[estado_idx]).strip()
        
        # fecha_pago
        fp_idx = col_map.get('fecha_pago')
        if fp_idx is not None and fp_idx < len(row) and row[fp_idx]:
            fp = row[fp_idx]
            if isinstance(fp, datetime):
                service['fecha_pago'] = fp
                service['estado_pago'] = 'pagado'
            elif isinstance(fp, str) and fp.strip():
                service['fecha_pago'] = fp.strip()
                service['estado_pago'] = 'pagado'
        
        # forma_pago
        fpago_idx = col_map.get('forma_pago')
        if fpago_idx is not None and fpago_idx < len(row) and row[fpago_idx]:
            service['forma_pago'] = str(row[fpago_idx]).strip()
        
        # descripcion
        desc_idx = col_map.get('descripcion')
        if desc_idx is not None and desc_idx < len(row) and row[desc_idx]:
            service['descripcion_residuo'] = str(row[desc_idx]).strip()
        
        services.append(service)
    
    return services

# Collect all service rows
all_services = []  # List of tuples for the Servicio sheet
service_id = 1
matched_count = 0
skipped_no_match = 0
skipped_no_sede = 0
skipped_no_header = 0
total_services_added = 0

for id_emp in sorted(empresa_ruc.keys()):
    ruc = empresa_ruc[id_emp]
    
    if ruc not in ruc_to_sheet:
        skipped_no_match += 1
        continue
    
    sheet_name = ruc_to_sheet[ruc]
    
    # Get sedes for this empresa
    sedes = empresa_sedes.get(id_emp)
    if not sedes:
        skipped_no_sede += 1
        continue
    
    # For each empresa, we use the first sede (most have only 1)
    # If multiple sedes, each sede gets the same services (as they come from the same PAGOS sheet)
    # Per the user's instructions, the services correspond to the sede
    # Since each PAGOS sheet is for one RUC (one empresa), and most have 1 sede,
    # we assign to the first sede
    id_sede = sedes[0][0]
    id_contrato = sede_contrato.get(id_sede)
    
    # Extract services from the PAGOS sheet
    ws = wb_pagos[sheet_name]
    header_idx, header_row = find_header_row(ws)
    
    if header_idx is None:
        skipped_no_header += 1
        continue
    
    services = extract_services(ws, header_idx)
    matched_count += 1
    
    for svc in services:
        row_data = (
            service_id,              # id_servicio
            id_emp,                  # id_empresa
            id_sede,                 # id_sede
            None,                    # id_ruta
            1,                       # id_planta (always "1")
            id_contrato,             # id_contrato
            svc['mes_servicio'],     # mes_servicio
            svc['fecha_ejecucion'],  # fecha_ejecucion
            svc['estado'],           # estado

            svc['estado_pago'],      # estado_pago
            svc['fecha_pago'],       # fecha_pago
            svc['forma_pago'],       # forma_pago
            svc['descripcion_residuo'],  # descripcion_residuo
            None,                    # fecha_creacion
            None,                    # fecha_modificacion
        )
        all_services.append(row_data)
        service_id += 1
    
    total_services_added += len(services)

wb_pagos.close()

print(f"\n--- Results ---")
print(f"Matched empresas: {matched_count}")
print(f"Skipped (no RUC match): {skipped_no_match}")
print(f"Skipped (no sede): {skipped_no_sede}")
print(f"Skipped (no header found): {skipped_no_header}")
print(f"Total service rows to write: {total_services_added}")

# --- Step 4: Write to DATA 2026.xlsx ---
print(f"\nCreating backup...")
backup_file = DATA_FILE.replace('.xlsx', '_BACKUP.xlsx')
shutil.copy2(DATA_FILE, backup_file)
print(f"  Backup saved to: {backup_file}")

print(f"\nWriting to Servicio sheet...")
wb_write = openpyxl.load_workbook(DATA_FILE)
ws_servicio = wb_write['Servicio']

# Clear existing data rows (keep header at row 1)
# Delete from row 2 to max_row
if ws_servicio.max_row > 1:
    ws_servicio.delete_rows(2, ws_servicio.max_row - 1)

# Write all service rows
for i, row_data in enumerate(all_services):
    row_num = i + 2  # Row 1 is header
    for col, value in enumerate(row_data, 1):
        ws_servicio.cell(row=row_num, column=col, value=value)

wb_write.save(DATA_FILE)
wb_write.close()

print(f"\nDone! Wrote {len(all_services)} service rows to Servicio sheet.")
print(f"First 5 rows (preview):")
for i, row in enumerate(all_services[:5]):
    print(f"  {row}")
