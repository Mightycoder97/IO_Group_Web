
import openpyxl
import re
from datetime import datetime, timedelta

# --- Configuración ---
INPUT_FILE = 'DATA 2026.xlsx'
SHEET_NAME = 'Data Completa'
OUTPUT_SQL = 'control/database/migrations/import_data_completa_2026.sql'

print(f"Cargando {INPUT_FILE}...")
wb = openpyxl.load_workbook(INPUT_FILE, data_only=True, read_only=True)

if SHEET_NAME not in wb.sheetnames:
    print(f"Error: Hoja '{SHEET_NAME}' no encontrada.")
    exit(1)

ws = wb[SHEET_NAME]
rows = list(ws.iter_rows(values_only=True))
headers = [str(h).strip().upper() if h else '' for h in rows[0]]

# Mapeo de columnas por nombre exacto (según lectura previa)
COL_MAP = {
    'REPRESENTANTE': -1,
    'DNI': -1,
    'STATUS': -1,
    'RAZON SOCIAL': -1,
    'RUC': -1,
    'RUBRO': -1,
    'DIRECCION': -1,
    'DISTRITO': -1,
    'PROVINCIA': -1,
    'DEPARTAMENTO': -1,
    'TELEFONO PARA COBRAR': -1,
    'NOMBRE COMERCIAL': -1,
    'CONTACTO': -1,
    'TELEFONO PARA PROGRAMAR': -1,
    'FECHA INICIO CONTRATO': -1,
    'LIMITE DE PESO': -1,
    'FRECUENCIA': -1,
    'TARIFA': -1
}

# Encontrar índices
for col_name in COL_MAP.keys():
    try:
        COL_MAP[col_name] = headers.index(col_name)
    except ValueError:
        print(f"Advertencia: Columna '{col_name}' no encontrada en headers exactos.")
        # Intentar búsqueda aprox
        for i, h in enumerate(headers):
            if col_name in h:
                COL_MAP[col_name] = i
                print(f"  -> Asignada a '{h}'")
                break

# Validar columnas críticas
if COL_MAP['RUC'] == -1:
    print("Error Crítico: No se encontró columna RUC")
    exit(1)

# Estructuras para evitar duplicados en el script
clientes_procesados = {} # nombre -> id_variable_sql
empresas_procesadas = {} # ruc -> id_variable_sql

sql_statements = []
sql_statements.append("-- Importación Data Completa 2026 (Cliente, Empresa, Sede, Contrato)")
sql_statements.append(f"-- Generado: {datetime.now()}")
sql_statements.append("SET FOREIGN_KEY_CHECKS = 0;")
sql_statements.append("SET @id_cliente_actual = 0;")
sql_statements.append("SET @id_empresa_actual = 0;")
sql_statements.append("SET @id_sede_actual = 0;")
sql_statements.append("")

def clean_text(val):
    if val is None: return ''
    return str(val).strip().replace("'", "\\'")

def get_status_activo(val):
    if val and str(val).strip().upper() == 'INACTIVO':
        return 0
    return 1

def parse_date(val):
    # Retorna string 'YYYY-MM-DD' o None
    if val is None: return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    
    val_str = str(val).strip()
    # Intentar formatos comunes
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%dT%H:%M:%S'):
        try:
            dt = datetime.strptime(val_str.split('T')[0], fmt.replace('T%H:%M:%S', ''))
            # Validar años razonables
            if dt.year < 2000 or dt.year > 2100:
                return None
            return dt.strftime('%Y-%m-%d')
        except:
            pass
    return None

def clean_decimal(val):
    if val is None: return '0.00'
    val_str = str(val).upper().replace('KG', '').replace('SIN LIMITE', '').strip()
    if not val_str: return 'NULL'
    try:
        # Extraer solo numeros y punto
        nums = re.findall(r"[-+]?\d*\.\d+|\d+", val_str)
        if nums:
            return nums[0]
        return 'NULL'
    except:
        return 'NULL'

count_rows = 0

for row_idx, row in enumerate(rows[1:], start=2):
    # Leer datos básicos
    ruc = clean_text(row[COL_MAP['RUC']]).replace(' ', '')
    if not ruc:
        continue # Sin RUC no hay empresa

    representante = clean_text(row[COL_MAP['REPRESENTANTE']])
    dni = clean_text(row[COL_MAP['DNI']])
    status_raw = row[COL_MAP['STATUS']] if COL_MAP['STATUS'] != -1 else 'ACTIVO'
    activo = get_status_activo(status_raw)
    
    razon_social = clean_text(row[COL_MAP['RAZON SOCIAL']])
    rubro = clean_text(row[COL_MAP['RUBRO']])
    direccion_fiscal = clean_text(row[COL_MAP['DIRECCION']])
    distrito = clean_text(row[COL_MAP['DISTRITO']])
    provincia = clean_text(row[COL_MAP['PROVINCIA']])
    departamento = clean_text(row[COL_MAP['DEPARTAMENTO']])
    # tel_cobranza = clean_text(row[COL_MAP['TELEFONO PARA COBRAR']]) -- REMOVIDO: No existe en tabla Empresa
    
    nombre_comercial = clean_text(row[COL_MAP['NOMBRE COMERCIAL']])
    if not nombre_comercial: nombre_comercial = razon_social
    
    contacto_sede = clean_text(row[COL_MAP['CONTACTO']])
    tel_programar = clean_text(row[COL_MAP['TELEFONO PARA PROGRAMAR']])
    
    fecha_inicio = parse_date(row[COL_MAP['FECHA INICIO CONTRATO']])
    peso_limite = clean_decimal(row[COL_MAP['LIMITE DE PESO']])
    frecuencia = clean_text(row[COL_MAP['FRECUENCIA']]).lower()
    tarifa = clean_decimal(row[COL_MAP['TARIFA']])
    if tarifa == 'NULL': tarifa = '0.00'

    # Calcular Fecha Fin (+365 días)
    fecha_fin = 'NULL'
    if fecha_inicio:
        try:
            dt_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            dt_fin = dt_inicio + timedelta(days=365)
            fecha_fin = f"'{dt_fin.strftime('%Y-%m-%d')}'"
            fecha_inicio = f"'{fecha_inicio}'"
        except:
             fecha_inicio = 'NULL'
    else:
        fecha_inicio = 'NULL'

    # Lógica Cliente
    # Determinar tipo doc cliente
    tipo_doc_cliente = 'Otro'
    if len(dni) == 8 and dni.isdigit():
        tipo_doc_cliente = 'DNI'
    elif len(dni) == 7 and dni.isdigit():
        dni = '0' + dni
        tipo_doc_cliente = 'DNI'
    else:
        tipo_doc_cliente = 'Carnet de Extranjeria'

    nombre_cliente = representante if representante else f"Cliente {razon_social}"[:100]

    sql_statements.append(f"-- Fila {row_idx}: {razon_social} ({ruc})")
    
    # 1. CLIENTE (Verificar existencia por nombre)
    if nombre_cliente not in clientes_procesados:
        sql_statements.append(f"""
INSERT INTO Cliente (nombre, dni, tipo_documento, activo)
SELECT '{nombre_cliente}', '{dni}', '{tipo_doc_cliente}', {activo}
WHERE NOT EXISTS (SELECT 1 FROM Cliente WHERE nombre = '{nombre_cliente}');
""")
        # Obtener ID
        sql_statements.append(f"SET @id_cliente_actual = (SELECT id_cliente FROM Cliente WHERE nombre = '{nombre_cliente}' LIMIT 1);")
        clientes_procesados[nombre_cliente] = True
    else:
        sql_statements.append(f"SET @id_cliente_actual = (SELECT id_cliente FROM Cliente WHERE nombre = '{nombre_cliente}' LIMIT 1);")

    # 2. EMPRESA
    # Insertar o Actualizar
    # OMITIDOS: rubro, telefono (no en schema backup)
    sql_statements.append(f"""
INSERT INTO Empresa (id_cliente, razon_social, ruc, direccion_fiscal, distrito, provincia, departamento, rubro, activo)
VALUES (@id_cliente_actual, '{razon_social}', '{ruc}', '{direccion_fiscal}', '{distrito}', '{provincia}', '{departamento}', '{rubro}', {activo})
ON DUPLICATE KEY UPDATE 
    id_cliente = @id_cliente_actual,
    razon_social = '{razon_social}',
    direccion_fiscal = '{direccion_fiscal}',
    distrito = '{distrito}',
    provincia = '{provincia}',
    departamento = '{departamento}',
    rubro = '{rubro}',
    activo = {activo};
""")
    sql_statements.append(f"SET @id_empresa_actual = (SELECT id_empresa FROM Empresa WHERE ruc = '{ruc}' LIMIT 1);")

    # 3. SEDE
    # Buscar ID de sede por nombre comercial, empresa y direccion (para distinguir sedes con mismo nombre)
    sql_statements.append(f"""
INSERT INTO Sede (id_empresa, nombre_comercial, direccion, distrito, provincia, departamento, contacto_nombre, contacto_telefono, activo)
SELECT @id_empresa_actual, '{nombre_comercial}', '{direccion_fiscal}', '{distrito}', '{provincia}', '{departamento}', '{contacto_sede}', '{tel_programar}', {activo}
WHERE NOT EXISTS (SELECT 1 FROM Sede WHERE id_empresa = @id_empresa_actual AND nombre_comercial = '{nombre_comercial}' AND direccion = '{direccion_fiscal}');
""")
    sql_statements.append(f"SET @id_sede_actual = (SELECT id_sede FROM Sede WHERE id_empresa = @id_empresa_actual AND nombre_comercial = '{nombre_comercial}' AND direccion = '{direccion_fiscal}' LIMIT 1);")

    # 4. CONTRATO
    freq_map = {
        'diario': 'diario', 'semanal': 'semanal', 'quincenal': 'quincenal', 
        'mensual': 'mensual', 'bimestral': 'bimestral', 'trimestral': 'trimestral', 'eventual': 'eventual'
    }
    freq_db = freq_map.get(frecuencia, 'mensual') # Default mensual if unknown

    if fecha_inicio != 'NULL':
        # Insertar nuevo si no existe para esta sede
        sql_statements.append(f"""
INSERT INTO ContratoServicio (id_sede, fecha_inicio, fecha_fin, frecuencia, peso_limite_kg, tarifa, activo)
SELECT @id_sede_actual, {fecha_inicio}, {fecha_fin}, '{freq_db}', {peso_limite}, {tarifa}, {activo}
WHERE NOT EXISTS (SELECT 1 FROM ContratoServicio WHERE id_sede = @id_sede_actual);
""")
        # Update si existe (para mantener actualizado contrato único por sede si ese es el modelo)
        sql_statements.append(f"""
UPDATE ContratoServicio SET 
    fecha_inicio = {fecha_inicio}, 
    fecha_fin = {fecha_fin}, 
    frecuencia = '{freq_db}', 
    peso_limite_kg = {peso_limite}, 
    tarifa = {tarifa}, 
    activo = {activo}
WHERE id_sede = @id_sede_actual;
""")

    count_rows += 1

sql_statements.append("SET FOREIGN_KEY_CHECKS = 1;")
sql_statements.append(f"-- Total procesados: {count_rows}")

with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_statements))

print(f"✅ SQL generado en: {OUTPUT_SQL}")
print(f"Total filas procesadas: {count_rows}")
