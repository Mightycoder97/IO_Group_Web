import pymysql
import re
import os
import openpyxl

db_config = {
    'host': '31.97.208.99',
    'port': 3306,
    'user': 'u511863531_Sebastian',
    'password': 'Sebas0920%',
    'database': 'u511863531_IOGroupBD',
    'charset': 'utf8mb4'
}

def clean_name(name):
    if not name: return ""
    name = str(name).lower()
    replacements = {'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u', 'ü': 'u', 'ñ': 'n'}
    for k, v in replacements.items():
        name = name.replace(k, v)
    name = re.sub(r'[.,\-–_/()\'"”’]', ' ', name)
    suffixes = [r'\bs\s*a\s*c\b', r'\be\s*i\s*r\s*l\b', r'\bs\s*a\b', r'\bs\s*r\s*l\b', r'\bcoop\b', r'\bcia\b', r'\bcompany\b']
    for s in suffixes:
        name = re.sub(s, '', name)
    return " ".join(name.split())

def extract_ruc_and_name_from_sheet(sheet):
    ruc = None
    client_name = None
    rows = list(sheet.iter_rows(values_only=True))
    if not rows: return None, None
    for r in rows[:12]:
        row_str = " | ".join([str(cell) for cell in r if cell is not None])
        if not ruc:
            ruc_match = re.search(r'\b(10|20)\d{9}\b', row_str)
            if ruc_match: ruc = ruc_match.group(0)
        if not client_name and any(cell is not None for cell in r):
            non_empty = [str(cell).strip() for cell in r if cell is not None]
            if non_empty: client_name = non_empty[0]
    return ruc, client_name

try:
    connection = pymysql.connect(**db_config)
    with connection.cursor(pymysql.cursors.DictCursor) as cursor:
        cursor.execute("SELECT id_empresa, ruc, razon_social FROM Empresa")
        db_empresas = cursor.fetchall()
        db_ruc_map = {row['ruc'].strip(): row for row in db_empresas if row['ruc']}
        
        cursor.execute("SELECT id_sede, id_empresa, nombre_comercial, direccion, distrito, provincia FROM Sede")
        db_sedes = cursor.fetchall()
        
    connection.close()
    
    excel_files = [
        "/Users/mightycoder/Documents/GitHub/IO_Group_Web/excel_data/PAGOS CLIENTES LIMA.xlsx",
        "/Users/mightycoder/Documents/GitHub/IO_Group_Web/excel_data/PAGOS CLIENTES SUR.xlsx"
    ]
    
    for filepath in excel_files:
        if not os.path.exists(filepath): continue
        wb = openpyxl.load_workbook(filepath, read_only=True)
        print(f"\nMatches for {os.path.basename(filepath)}:")
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            ruc, first_row_name = extract_ruc_and_name_from_sheet(sheet)
            if ruc: ruc = ruc.strip()
            
            emp_id = None
            mtype = None
            
            if ruc and ruc in db_ruc_map:
                emp_id = db_ruc_map[ruc]['id_empresa']
                mtype = "RUC Exact"
            else:
                s_name_clean = clean_name(sheetname)
                fr_name_clean = clean_name(first_row_name)
                
                # Check exact razon social
                for emp in db_empresas:
                    rs_clean = clean_name(emp['razon_social'])
                    if s_name_clean == rs_clean or (fr_name_clean and fr_name_clean == rs_clean):
                        emp_id = emp['id_empresa']
                        mtype = "Name Exact (Razon Social)"
                        break
                        
                if not emp_id:
                    # Check exact nombre comercial
                    for sd in db_sedes:
                        nc_clean = clean_name(sd['nombre_comercial'])
                        if s_name_clean == nc_clean or (fr_name_clean and fr_name_clean == nc_clean):
                            emp_id = sd['id_empresa']
                            mtype = "Name Exact (Nombre Comercial)"
                            break
                            
                if not emp_id:
                    # Check substring razon social
                    for emp in db_empresas:
                        rs_clean = clean_name(emp['razon_social'])
                        if (s_name_clean and len(s_name_clean) >= 4 and s_name_clean in rs_clean) or \
                           (rs_clean and len(rs_clean) >= 4 and rs_clean in s_name_clean) or \
                           (fr_name_clean and len(fr_name_clean) >= 4 and fr_name_clean in rs_clean) or \
                           (rs_clean and len(rs_clean) >= 4 and rs_clean in fr_name_clean):
                            emp_id = emp['id_empresa']
                            mtype = "Substring (Razon Social)"
                            break
                            
                if not emp_id:
                    # Check substring nombre comercial
                    for sd in db_sedes:
                        nc_clean = clean_name(sd['nombre_comercial'])
                        if (s_name_clean and len(s_name_clean) >= 4 and s_name_clean in nc_clean) or \
                           (nc_clean and len(nc_clean) >= 4 and nc_clean in s_name_clean) or \
                           (fr_name_clean and len(fr_name_clean) >= 4 and fr_name_clean in nc_clean) or \
                           (nc_clean and len(nc_clean) >= 4 and nc_clean in fr_name_clean):
                            emp_id = sd['id_empresa']
                            mtype = "Substring (Nombre Comercial)"
                            break
                            
            if emp_id:
                # Find company name
                comp_name = next((e['razon_social'] for e in db_empresas if e['id_empresa'] == emp_id), "Unknown")
                # Let's print matches of interest
                targets = ["PETMAX", "MARTINEZ", "CURI", "CLINIVET", "LASSY"]
                if any(t in sheetname.upper() or t in comp_name.upper() for t in targets):
                    print(f"  Sheet: '{sheetname}' -> Company: '{comp_name}' (ID: {emp_id}) via {mtype}")
                    
        wb.close()
except Exception as e:
    print(f"Error: {e}")
