import openpyxl
import os

excel_path = "/Users/mightycoder/Documents/GitHub/IO_Group_Web/excel_data/PAGOS CLIENTES LIMA.xlsx"

if os.path.exists(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print("Searching for 7772...")
    for sname in wb.sheetnames:
        sheet = wb[sname]
        for r in range(1, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                val = sheet.cell(row=r, column=c).value
                if val is not None and ('7772' in str(val) or val == 7772 or val == 7772.0):
                    print(f"  Match in Sheet '{sname}' | Cell ({r}, {c}): {val}")
    wb.close()
else:
    print("Excel file not found!")
