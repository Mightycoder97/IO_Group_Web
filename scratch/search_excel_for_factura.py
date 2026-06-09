import openpyxl
import os

excel_dir = "/Users/mightycoder/Documents/GitHub/IO_Group_Web/excel_data"
target_fact = "42379"

for fname in ["PAGOS CLIENTES LIMA.xlsx", "PAGOS CLIENTES SUR.xlsx"]:
    fpath = os.path.join(excel_dir, fname)
    if os.path.exists(fpath):
        wb = openpyxl.load_workbook(fpath, data_only=True)
        print(f"\nSearching for Factura {target_fact} in {fname}...")
        for sname in wb.sheetnames:
            sheet = wb[sname]
            for r in range(1, sheet.max_row + 1):
                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=r, column=c).value
                    if val is not None and (target_fact in str(val)):
                        print(f"  Match in Sheet '{sname}' | Cell ({r}, {c}): {val} | Full Row: {[sheet.cell(row=r, column=col).value for col in range(1, 13)]}")
        wb.close()
