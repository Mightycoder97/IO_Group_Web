import openpyxl
import os

excel_path = "/Users/mightycoder/Documents/GitHub/IO_Group_Web/excel_data/PAGOS CLIENTES SUR.xlsx"

if os.path.exists(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Find matching sheet
    matching_sheets = [s for s in wb.sheetnames if 'HERNANI' in s.upper()]
    print("Matching sheets in SUR:", matching_sheets)
    
    for sheet_name in matching_sheets:
        sheet = wb[sheet_name]
        print(f"\n--- Rows in SUR Sheet '{sheet_name}' ---")
        for r in range(1, 40):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
            if any(row_vals):
                print(f"  Row {r}: {row_vals}")
    wb.close()
else:
    print("Excel SUR file not found!")
