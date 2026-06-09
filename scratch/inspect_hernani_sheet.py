import openpyxl
import os

excel_path = "/Users/mightycoder/Documents/GitHub/IO_Group_Web/excel_data/PAGOS CLIENTES LIMA.xlsx"

if os.path.exists(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = 'CLINICA HERNANI SAC       '
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"--- Rows in sheet '{sheet_name}' ---")
        for r in range(1, 40):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
            if any(row_vals):
                print(f"  Row {r}: {row_vals}")
    wb.close()
else:
    print("Excel file not found!")
