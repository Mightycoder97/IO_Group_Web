import openpyxl
import os

excel_dir = "/Users/mightycoder/Documents/GitHub/IO_Group_Web/excel_data"
target_val = "247985"

for fname in ["PAGOS CLIENTES LIMA.xlsx", "PAGOS CLIENTES SUR.xlsx"]:
    fpath = os.path.join(excel_dir, fname)
    if os.path.exists(fpath):
        wb = openpyxl.load_workbook(fpath, data_only=True)
        print(f"\nSearching for {target_val} in {fname}...")
        for sname in wb.sheetnames:
            sheet = wb[sname]
            for r in range(1, sheet.max_row + 1):
                for c in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=r, column=c).value
                    if val is not None and (target_val in str(val)):
                        print(f"  Match in Sheet '{sname}' | Cell ({r}, {c}): {val}")
        wb.close()
