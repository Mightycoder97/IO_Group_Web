import openpyxl
import os
from datetime import datetime

EXCEL_FILE = 'DATA 2026.xlsx'
OUTPUT_SQL = 'control/database/import_data_2026.sql'

def clean_value(val):
    if val is None:
        return 'NULL'
    if isinstance(val, str):
        val = val.strip().replace("'", "\\'")
        if not val:
            return 'NULL'
        return f"'{val}'"
    if isinstance(val, (int, float)):
        return str(val)
    if isinstance(val, datetime):
        return f"'{val.strftime('%Y-%m-%d %H:%M:%S')}'"
    return f"'{str(val)}'"

def parse_estado(val):
    if not val:
        return "'programado'"
    
    s = str(val).strip().upper()
    
    # Operational Status Mapping
    # Keywords indicating the service was ABORTED/NOT DONE
    if any(x in s for x in ['NO SE HIZO', 'NO UBICO', 'NO LE HICIERON', 'REPROGRAMADO', 'ANULADO', 'FALSO FLETE']):
        return "'cancelado'"
    
    # "CANCELADO" in accounting means PAID, so the service was COMPLETED.
    # Payment methods also imply completion.
    return "'completado'"

def generate_sql():
    print(f"Reading {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    
    with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
        f.write("-- Import script for DATA 2026\n")
        f.write("-- Generated automatically\n\n")
        f.write("SET FOREIGN_KEY_CHECKS = 0;\n\n")
        
        # 1. CLIENTES
        if 'CLIENTES' in wb.sheetnames:
            print("Processing CLIENTES...")
            ws = wb['CLIENTES']
            f.write("-- ==========================================\n")
            f.write("-- TABLA: Cliente\n")
            f.write("-- ==========================================\n")
            f.write("DELETE FROM Cliente;\n")
            f.write("ALTER TABLE Cliente AUTO_INCREMENT = 1;\n")
            
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for row in rows:
                if not row[0]: continue
                # id_cliente, nombre, tipo_documento, dni, activo
                # Schema: id_cliente, nombre, tipo_documento, dni, activo, fecha_creacion, ...
                val_id = clean_value(row[0])
                val_nombre = clean_value(row[1])
                val_tipo = clean_value(row[2]) if row[2] else "'DNI'"
                val_dni = clean_value(row[3])
                val_activo = clean_value(row[4]) if len(row) > 4 else '1'
                
                f.write(f"INSERT INTO Cliente (id_cliente, nombre, tipo_documento, dni, activo) VALUES ({val_id}, {val_nombre}, {val_tipo}, {val_dni}, {val_activo});\n")
            f.write("\n")

        # 2. EMPRESA
        if 'EMPRESA' in wb.sheetnames:
            print("Processing EMPRESA...")
            ws = wb['EMPRESA']
            f.write("-- ==========================================\n")
            f.write("-- TABLA: Empresa\n")
            f.write("-- ==========================================\n")
            f.write("DELETE FROM Empresa;\n")
            
            # Use SET to track seen RUCs
            seen_rucs = set()
            
            f.write("ALTER TABLE Empresa AUTO_INCREMENT = 1;\n")
            
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for row in rows:
                if not row[0]: continue
                # id_empresa, id_cliente, razon_social, ruc, ...
                val_ruc_raw = row[3]
                
                # Check for duplicates
                if val_ruc_raw in seen_rucs:
                    print(f"Skipping duplicate RUC: {val_ruc_raw}")
                    continue
                seen_rucs.add(val_ruc_raw)

                val_id = clean_value(row[0])
                val_cli = clean_value(row[1])
                val_rs = clean_value(row[2])
                val_ruc = clean_value(val_ruc_raw)
                val_dir = clean_value(row[4])
                val_dist = clean_value(row[5])
                val_prov = clean_value(row[6])
                val_dep = clean_value(row[7])
                val_act = clean_value(row[8]) if len(row) > 8 else '1'
                
                f.write(f"INSERT INTO Empresa (id_empresa, id_cliente, razon_social, ruc, direccion_fiscal, distrito, provincia, departamento, activo) VALUES ({val_id}, {val_cli}, {val_rs}, {val_ruc}, {val_dir}, {val_dist}, {val_prov}, {val_dep}, {val_act});\n")
            f.write("\n")

        # 3. SEDE
        if 'SEDE' in wb.sheetnames:
            print("Processing SEDE...")
            ws = wb['SEDE']
            f.write("-- ==========================================\n")
            f.write("-- TABLA: Sede\n")
            f.write("-- ==========================================\n")
            f.write("DELETE FROM Sede;\n")
            f.write("ALTER TABLE Sede AUTO_INCREMENT = 1;\n")
            
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for row in rows:
                if not row[0]: continue
                # id_sede, id_empresa, nombre_comercial, direccion, ...
                val_id = clean_value(row[0])
                val_emp = clean_value(row[1])
                val_nom = clean_value(row[2])
                val_dir = clean_value(row[3])
                val_dist = clean_value(row[4])
                val_prov = clean_value(row[5])
                val_dep = clean_value(row[6])
                val_ref = clean_value(row[7])
                val_gps = clean_value(row[8])
                val_cn = clean_value(row[9])
                val_ct = clean_value(row[10]) # contacto_telefono
                # row[11] is contacto_telefono_2, row[12] is _3 - skipping or merging? Schema has only one. 
                # Let's use just the first one for now as per schema
                val_ce = clean_value(row[13])
                val_act = clean_value(row[14]) if len(row) > 14 else '1'
                
                f.write(f"INSERT INTO Sede (id_sede, id_empresa, nombre_comercial, direccion, distrito, provincia, departamento, referencia, coordenadas_gps, contacto_nombre, contacto_telefono, contacto_email, activo) VALUES ({val_id}, {val_emp}, {val_nom}, {val_dir}, {val_dist}, {val_prov}, {val_dep}, {val_ref}, {val_gps}, {val_cn}, {val_ct}, {val_ce}, {val_act});\n")
            f.write("\n")

        # 4. ContratoServicio
        if 'ContratoServicio' in wb.sheetnames:
            print("Processing ContratoServicio...")
            ws = wb['ContratoServicio']
            f.write("-- ==========================================\n")
            f.write("-- TABLA: ContratoServicio\n")
            f.write("-- ==========================================\n")
            f.write("DELETE FROM ContratoServicio;\n")
            f.write("ALTER TABLE ContratoServicio AUTO_INCREMENT = 1;\n")
            
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for row in rows:
                if not row[0]: continue
                # id_contrato, id_sede, fecha_inicio, fecha_fin, frecuencia, peso_limite_kg, tarifa, tipo_tarifa, doc, obs
                val_id = clean_value(row[0])
                val_sede = clean_value(row[1])
                val_fi = clean_value(row[2])
                val_ff = clean_value(row[3])
                val_freq = clean_value(row[4])
                val_peso = clean_value(row[5])
                
                # Tarifa cannot be NULL. Default to 0.00 if missing.
                val_tarifa = clean_value(row[6])
                if val_tarifa == 'NULL':
                    val_tarifa = '0.00'
                
                val_tipo = clean_value(row[7]) # tipo_tarifa
                val_doc = clean_value(row[8])
                val_obs = clean_value(row[9]) # observaciones IS present in ContratoServicio
                
                f.write(f"INSERT INTO ContratoServicio (id_contrato, id_sede, fecha_inicio, fecha_fin, frecuencia, peso_limite_kg, tarifa, tipo_tarifa, doc_escaneado, observaciones) VALUES ({val_id}, {val_sede}, {val_fi}, {val_ff}, {val_freq}, {val_peso}, {val_tarifa}, {val_tipo}, {val_doc}, {val_obs});\n")
            f.write("\n")

        # 5. Servicio
        if 'Servicio' in wb.sheetnames:
            print("Processing Servicio...")
            ws = wb['Servicio']
            f.write("-- ==========================================\n")
            f.write("-- TABLA: Servicio\n")
            f.write("-- ==========================================\n")
            f.write("DELETE FROM Servicio;\n")
            f.write("ALTER TABLE Servicio AUTO_INCREMENT = 1;\n")
            
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            for row in rows:
                if not row[0]: continue
                # id_servicio, id_empresa, id_sede, id_ruta, id_planta, id_contrato, mes_servicio, 
                # fecha_ejecucion, estado, observaciones, estado_pago, fecha_pago, forma_pago, descripcion_residuo, ...
                
                val_id = clean_value(row[0])
                # id_empresa at row[1] is NOT in Servicio table directly (it's via Sede), but let's check schema.
                # Schema: id_servicio, id_sede, id_ruta, id_planta, id_contrato
                
                val_sede = clean_value(row[2])
                val_ruta = clean_value(row[3])
                val_planta = clean_value(row[4])
                val_contrato = clean_value(row[5])
                val_mes = clean_value(row[6])
                val_fej = clean_value(row[7])
                
                # Parse operational state from Excel 'estado' column
                val_estado = parse_estado(row[8])
                
                # Correct Mapping based on populate_servicio.py (Excel structure)
                # row[0]=id, [1]=emp(skip), [2]=sede, [3]=ruta, [4]=planta, [5]=contrato, [6]=mes, [7]=fecha_ej, [8]=estado
                # row[9]=estado_pago, [10]=fecha_pago, [11]=forma_pago, [12]=descripcion_residuo
                
                val_ep = clean_value(row[9])  # estado_pago
                
                # Logic: If state is 'completado' (meaning it was PAID/CANCELADO), 
                # and estado_pago is NULL/empty, default it to 'pagado'
                if val_ep == 'NULL' and val_estado == "'completado'":
                    val_ep = "'pagado'"
                elif val_ep == 'NULL':
                    val_ep = "'pendiente'"
                
                val_fp = clean_value(row[10]) # fecha_pago
                val_forma = clean_value(row[11]) # forma_pago
                val_desc = clean_value(row[12]) # descripcion_residuo
                
                f.write(f"INSERT INTO Servicio (id_servicio, id_sede, id_ruta, id_planta, id_contrato, mes_servicio, fecha_ejecucion, estado, estado_pago, fecha_pago, forma_pago, descripcion_residuo) VALUES ({val_id}, {val_sede}, {val_ruta}, {val_planta}, {val_contrato}, {val_mes}, {val_fej}, {val_estado}, {val_ep}, {val_fp}, {val_forma}, {val_desc});\n")
            f.write("\n")
            
        f.write("SET FOREIGN_KEY_CHECKS = 1;\n")
            
    print(f"Generated {OUTPUT_SQL}")

if __name__ == '__main__':
    generate_sql()
