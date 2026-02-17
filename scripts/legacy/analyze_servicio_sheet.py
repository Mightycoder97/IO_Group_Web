
import openpyxl

INPUT_FILE = 'DATA 2026.xlsx'
SHEET_NAME = 'Servicio'

print(f"Leyendo '{SHEET_NAME}' en '{INPUT_FILE}'...")
try:
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Error: Hoja '{SHEET_NAME}' no encontrada.")
        exit()
    
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True, max_row=5))
    
    headers = [str(h).strip() if h else f"COL_{i}" for i, h in enumerate(rows[0])]
    print(f"Encabezados: {headers}")
    
    print("\nEjemplo de datos:")
    for i, row in enumerate(rows[1:]):
        print(f"Fila {i+1}: {row}")
        
except Exception as e:
    print(f"Error: {e}")
