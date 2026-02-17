
import openpyxl

INPUT_FILE = 'PAGOS CLIENTES LIMA.xlsx'
SHEETS_TO_DEBUG = ['CLINICA VETERINARIA SAN ROQUE L', 'ROSA MILAGRITOS MARTINEZ BRAVO']

print(f"Cargando {INPUT_FILE}...")
wb = openpyxl.load_workbook(INPUT_FILE, data_only=True, read_only=True)

for s_name in SHEETS_TO_DEBUG:
    if s_name not in wb.sheetnames:
        print(f"Hoja '{s_name}' no encontrada.")
        continue

    print(f"\n--- DEBUG: {s_name} ---")
    ws = wb[s_name]
    rows = list(ws.iter_rows(values_only=True, max_row=25))
    
    for i, row in enumerate(rows):
        # Print non-empty rows
        if any(row):
            print(f"Row {i+1}: {row}")
