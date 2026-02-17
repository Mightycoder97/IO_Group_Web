
import openpyxl

INPUT_FILE = 'PAGOS CLIENTES LIMA.xlsx'

print(f"Leyendo archivo '{INPUT_FILE}'...")
try:
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    print(f"Hojas encontradas: {sheet_names}")
    
    # Read first sheet by default or 'PAGOS CLIENTES LIMA' if exists
    target_sheet = sheet_names[0]
    for n in sheet_names:
        if 'PAGO' in n.upper():
            target_sheet = n
            break
            
    print(f"Analizando hoja: {target_sheet}")
    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True, max_row=10))
    
    headers = [str(h).strip() if h else f"COL_{i}" for i, h in enumerate(rows[0])]
    print(f"Encabezados: {headers}")
    
    print("\nEjemplo de datos (primeras 3 filas):")
    for i, row in enumerate(rows[1:4]):
        print(f"Fila {i+1}: {row}")
        
except Exception as e:
    print(f"Error: {e}")
