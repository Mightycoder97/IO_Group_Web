
import openpyxl
import re

INPUT_FILE = 'PAGOS CLIENTES LIMA.xlsx'

print(f"Leyendo archivo '{INPUT_FILE}'...")
try:
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    
    # Pick a few sheets: First, Middle, Last, and some random ones
    indices = [0, 10, 100, 500, len(sheet_names)-1]
    sheets_to_check = [sheet_names[i] for i in indices if i < len(sheet_names)]

    for s_name in sheets_to_check:
        print(f"\n--- Analizando Hoja: {s_name} ---")
        ws = wb[s_name]
        rows = list(ws.iter_rows(values_only=True, max_row=15))
        
        ruc_found = None
        header_row_idx = -1
        
        for i, row in enumerate(rows):
            row_str = " ".join([str(c) for c in row if c]).upper()
            
            # Buscando RUC
            if 'RUC' in row_str and not ruc_found:
                # Try to extract 11 digit number
                match = re.search(r'\b(10|20)\d{9}\b', row_str)
                if match:
                    ruc_found = match.group(0)
                    print(f"RUC Encontrado en fila {i+1}: {ruc_found}")
                else: 
                     # look in cells
                     for cell in row:
                         if isinstance(cell, (int, str)) and re.match(r'^(10|20)\d{9}$', str(cell)):
                             ruc_found = str(cell)
                             print(f"RUC Encontrado en celda fila {i+1}: {ruc_found}")
                             break
            
            # Buscando Encabezado (keywords comunes)
            if header_row_idx == -1 and 'FECHA' in row_str and ('SERVICIO' in row_str or 'PAGO' in row_str):
                header_row_idx = i
                print(f"Encabezado detectado en fila {i+1}: {row}")

        if not ruc_found:
            print("⚠️ No se encontró RUC en las primeras 15 filas.")
        if header_row_idx == -1:
            print("⚠️ No se detectó fila de encabezado clara.")

except Exception as e:
    print(f"Error: {e}")
