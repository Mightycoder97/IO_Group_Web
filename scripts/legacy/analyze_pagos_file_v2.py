
import openpyxl

INPUT_FILE = 'PAGOS CLIENTES LIMA.xlsx'

print(f"Leyendo archivo '{INPUT_FILE}'...")
try:
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    print(f"Total hojas: {len(sheet_names)}")
    print(f"Primeras 10 hojas: {sheet_names[:10]}")
    
    # Check for likely summary sheets
    candidates = ['RESUMEN', 'GENERAL', 'DATA', 'BASE', 'PAGOS', 'LIMA', 'CONSOLIDADO', 'AMBOS', 'PAGOS 2026', 'PAGOS 2025']
    target_sheet = None
    for cand in candidates:
        for name in sheet_names:
            if cand in name.upper():
                target_sheet = name
                break
        if target_sheet: break
    
    if not target_sheet:
        print("No se encontró hoja resumen obvia. Analizando las primeras 3 hojas para ver estructura.")
        sheets_to_check = sheet_names[:3]
    else:
        print(f"Hoja candidata encontrada: {target_sheet}")
        sheets_to_check = [target_sheet]

    for s_name in sheets_to_check:
        print(f"\n--- Analizando Hoja: {s_name} ---")
        ws = wb[s_name]
        rows = list(ws.iter_rows(values_only=True, max_row=10))
        
        # Print first few non-empty rows to find header
        for i, row in enumerate(rows):
            # Check if row has at least 3 non-null values to be considered potential header
            non_null = [x for x in row if x]
            if len(non_null) > 3:
                print(f"Posible encabezado (Fila {i+1}): {row}")
                # Print next row as data sample
                if i + 1 < len(rows):
                    print(f"Datos (Fila {i+2}): {rows[i+1]}")
                break
        else:
            print("No se detectó fila de encabezado clara.")

except Exception as e:
    print(f"Error: {e}")
