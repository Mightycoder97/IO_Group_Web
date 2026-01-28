"""
Script para actualizar teléfonos y nombres de contacto desde Datos.xlsx
Extrae números limpios y nombres de contacto de formatos complejos
Genera SQL de actualización para la tabla Sede
"""
import openpyxl
import re
from datetime import datetime


def extract_phones_and_contacts(text):
    """
    Extrae TODOS los teléfonos y nombres de contacto de un texto.
    Retorna una lista de diccionarios: [{'phone': '945902609', 'name': 'Monica recp'}, ...]
    
    Maneja formatos como:
    - 937144263
    - 945 902 609 - Monica recp
    - 940709263/983726186
    - +51 998 170 294
    - 01 2635746
    - NUEVO NUMERO 902631960. 955 196 765
    - 996153628 - dra claudia
    """
    if not text or str(text).strip() == '':
        return []
    
    text = str(text)
    results = []
    
    # Remover prefijos de país
    text = text.replace('+51', '').replace('+', '')
    
    # Dividir por separadores comunes primero
    # Separadores: / , puntos cuando hay espacios alrededor
    segments = re.split(r'[/,]|\.\s+|\s{2,}', text)
    
    for segment in segments:
        segment = segment.strip()
        if not segment:
            continue
        
        # Patrón para capturar teléfono seguido opcionalmente de un nombre
        # Busca: dígitos con espacios (6-11 chars) + separador opcional + nombre opcional
        pattern = r'(\d[\d\s]{4,10}\d)\s*[-–]?\s*([a-zA-ZáéíóúñÁÉÍÓÚÑ\s\.]+)?'
        
        match = re.search(pattern, segment)
        
        if match:
            phone_raw = match.group(1)
            name_raw = match.group(2) if match.group(2) else None
            
            # Limpiar teléfono
            phone = re.sub(r'\D', '', phone_raw)
            
            # Validar longitud (teléfonos peruanos: 6-9 dígitos, con prefijo 01/51: hasta 11)
            if not (6 <= len(phone) <= 11):
                continue
            
            # Limpiar nombre
            name = None
            if name_raw:
                name = name_raw.strip()
                # Ignorar si es muy corto o parece basura
                if len(name) < 3:
                    name = None
                # Ignorar palabras clave que no son nombres
                ignore_words = ['numero', 'nuevo', 'tel', 'fono', 'cel', 'mov', 'fijo', 'programar']
                if name and name.lower() in ignore_words:
                    name = None
            
            # Evitar duplicados
            if not any(r['phone'] == phone for r in results):
                results.append({'phone': phone, 'name': name})
    
    # Si no encontró nada, buscar números simples
    if not results:
        simple_numbers = re.findall(r'\d{6,11}', re.sub(r'\s', '', text))
        for num in simple_numbers:
            if 6 <= len(num) <= 11 and not any(r['phone'] == num for r in results):
                results.append({'phone': num, 'name': None})
    
    return results


def main(test_mode=False):
    print("=" * 50)
    print("Actualizador de Teléfonos y Contactos")
    print("=" * 50)
    print()
    
    print("Cargando Datos.xlsx...")
    wb = openpyxl.load_workbook('Datos.xlsx', data_only=True)
    ws = wb['Sedes']
    
    updates = []
    
    for row in range(2, ws.max_row + 1):
        ruc = ws.cell(row=row, column=5).value  # Columna RUC
        sede = ws.cell(row=row, column=6).value  # Columna SEDE
        tel1_raw = ws.cell(row=row, column=11).value  # Columna Telefono
        tel2_raw = ws.cell(row=row, column=12).value  # Columna Telefono 2
        
        if not ruc:
            continue
        
        ruc = str(ruc).strip()
        if len(ruc) != 11:
            continue
        
        # Extraer todos los teléfonos y contactos
        contacts1 = extract_phones_and_contacts(tel1_raw)
        contacts2 = extract_phones_and_contacts(tel2_raw)
        
        # Combinar todos los contactos
        all_contacts = contacts1 + contacts2
        
        if not all_contacts:
            continue
        
        # Reorganizar: primer teléfono siempre debe tener valor
        # Tomar hasta 2 teléfonos únicos
        unique_phones = []
        for c in all_contacts:
            if c['phone'] not in [p['phone'] for p in unique_phones]:
                unique_phones.append(c)
            if len(unique_phones) >= 2:
                break
        
        tel1 = unique_phones[0]['phone'] if len(unique_phones) > 0 else None
        name1 = unique_phones[0]['name'] if len(unique_phones) > 0 else None
        tel2 = unique_phones[1]['phone'] if len(unique_phones) > 1 else None
        name2 = unique_phones[1]['name'] if len(unique_phones) > 1 else None
        
        # Usar el primer nombre disponible para contacto_nombre
        contact_name = name1 or name2
        
        if tel1:
            updates.append({
                'ruc': ruc,
                'sede': sede,
                'tel1_raw': tel1_raw,
                'tel2_raw': tel2_raw,
                'tel1': tel1,
                'tel2': tel2,
                'contact_name': contact_name,
            })
    
    print(f"Total registros con teléfonos: {len(updates)}")
    print()
    
    if test_mode:
        print("=== MODO TEST: Primeros 30 registros ===")
        for i, u in enumerate(updates[:30]):
            print(f"\n[{i+1}] RUC: {u['ruc']} | Sede: {str(u['sede'])[:30]}")
            print(f"    Raw Tel1: {u['tel1_raw']}")
            print(f"    Raw Tel2: {u['tel2_raw']}")
            print(f"    -> Tel1: {u['tel1']} | Tel2: {u['tel2']} | Contacto: {u['contact_name']}")
        return
    
    # Generar SQL
    print("Generando SQL de actualización...")
    
    sql_lines = [
        "-- ============================================",
        "-- Actualización de Teléfonos y Nombres de Contacto",
        f"-- Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"-- Total registros: {len(updates)}",
        "-- ============================================",
        "",
        "-- NOTA: Ejecutar DESPUÉS de la migración add_contacto_telefono_2.sql",
        "",
    ]
    
    for u in updates:
        # Construir SET clause
        set_parts = []
        
        # Siempre hay tel1 (por la validación anterior)
        set_parts.append(f"contacto_telefono = '{u['tel1']}'")
        
        if u['tel2']:
            set_parts.append(f"contacto_telefono_2 = '{u['tel2']}'")
        
        if u['contact_name']:
            name_escaped = u['contact_name'].replace("'", "\\'")
            set_parts.append(f"contacto_nombre = '{name_escaped}'")
        
        set_clause = ", ".join(set_parts)
        sede_condition = ""
        if u['sede']:
            sede_name = str(u['sede']).replace("'", "\\'")[:100]
            sede_condition = f" AND s.nombre_comercial LIKE '%{sede_name}%'"
        
        sql = f"""UPDATE Sede s
INNER JOIN Empresa e ON s.id_empresa = e.id_empresa
SET {set_clause}
WHERE e.ruc = '{u['ruc']}'{sede_condition};"""
        sql_lines.append(sql)
        sql_lines.append("")
    
    # Guardar SQL
    output_file = 'control/database/migrations/update_telefonos.sql'
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(sql_lines))
    
    print(f"SQL guardado en: {output_file}")
    print()
    print("Próximos pasos:")
    print("1. Ejecutar add_contacto_telefono_2.sql en phpMyAdmin")
    print("2. Ejecutar update_telefonos.sql en phpMyAdmin")
    print("3. Verificar en la cartilla de un cliente")


if __name__ == '__main__':
    import sys
    test_mode = '--test' in sys.argv
    main(test_mode)
