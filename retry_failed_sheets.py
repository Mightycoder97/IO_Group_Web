
import openpyxl
import re
from datetime import datetime
import unidecode

# --- Configuración ---
INPUT_FILE = 'PAGOS CLIENTES LIMA.xlsx'
FAILED_LIST_FILE = 'manual_review_needed.txt'
OUTPUT_SQL = 'control/database/migrations/import_servicios_2026_retry.sql'

# --- Utilitarios ---
DISTRITOS = [
    'ANCON', 'ATE', 'BARRANCO', 'BREÑA', 'CARABAYLLO', 'CHACLACAYO', 'CHORRILLOS', 'CIENEGUILLA',
    'COMAS', 'EL AGUSTINO', 'INDEPENDENCIA', 'JESUS MARIA', 'LA MOLINA', 'LA VICTORIA', 'LIMA',
    'LINCE', 'LOS OLIVOS', 'LURIGANCHO', 'LURIN', 'MAGDALENA', 'MIRAFLORES', 'PACHACAMAC',
    'PUCUSANA', 'PUEBLO LIBRE', 'PUENTE PIEDRA', 'PUNTA HERMOSA', 'PUNTA NEGRA', 'RIMAC',
    'SAN BARTOLO', 'SAN BORJA', 'SAN ISIDRO', 'SAN JUAN DE LURIGANCHO', 'SAN JUAN DE MIRAFLORES',
    'SAN LUIS', 'SAN MARTIN DE PORRES', 'SAN MIGUEL', 'SANTA ANITA', 'SANTA MARIA DEL MAR',
    'SANTA ROSA', 'SANTIAGO DE SURCO', 'SURQUILLO', 'VILLA EL SALVADOR', 'VILLA MARIA DEL TRIUNFO'
]

def normalize_text(text):
    if not text: return ""
    return unidecode.unidecode(str(text)).upper().strip()

def clean_sql_val(val):
    if val is None: return 'NULL'
    val_str = str(val).replace("'", "\\'")
    return f"'{val_str}'"

def parse_date_excel(cell_val):
    if isinstance(cell_val, datetime):
        return cell_val.strftime('%Y-%m-%d')
    if not cell_val:
        return None
    val_str = str(cell_val).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y.%m.%d', '%d.%m.%Y', '%d/%m/%y', '%d-%m-%y'):
        try:
            dt = datetime.strptime(val_str.split(' ')[0], fmt)
            if 2000 < dt.year < 2100:
                return dt.strftime('%Y-%m-%d')
            # Handle 2-digit year adjustment if needed (python does this mostly likely 20xx)
            if 0 < dt.year < 100:
                 dt = dt.replace(year=2000+dt.year)
                 return dt.strftime('%Y-%m-%d')
        except:
            pass
    return None

def normalize_residuo(desc):
    desc = normalize_text(desc)
    if 'BIO' in desc: return 'Biocontaminado'
    if 'ESPECIAL' in desc: return 'Especial'
    if 'CADAVER' in desc or 'ANIMAL' in desc: return 'Cadaver Animal'
    if 'MEDICAMENTO' in desc or 'VENCIDO' in desc: return 'Medicamento Vencido'
    return 'Otros'

# --- Leer lista de fallidos ---
sheets_to_retry = []
try:
    with open(FAILED_LIST_FILE, 'r') as f:
        for line in f:
            # Extract sheet name from log line
            # Default format: "LEVEL: Hoja 'SHEET_NAME' msg..."
            match = re.search(r"Hoja '([^']+)'", line)
            if match:
                sheets_to_retry.append(match.group(1))
except FileNotFoundError:
    print(f"No se encontró {FAILED_LIST_FILE}")
    exit()

print(f"Reintentando {len(sheets_to_retry)} hojas fallidas...")

print(f"Cargando archivo '{INPUT_FILE}'...")
wb = openpyxl.load_workbook(INPUT_FILE, data_only=True, read_only=True)

sql_statements = []
sql_statements.append("-- Importación Retry (PAGOS CLIENTES LIMA)")
sql_statements.append("SET FOREIGN_KEY_CHECKS = 0;")
sql_statements.append("-- Asegurar existencia de Planta Principal (Requerido por FK)")
sql_statements.append("INSERT IGNORE INTO Planta (id_planta, razon_social, ruc, direccion, activo) VALUES (1, 'PLANTA PRINCIPAL', '20000000000', 'LIMA', 1);")

count_recovered = 0
count_services_generated = 0
recovered_sheets = []
failed_reasons = {}

for sheet_name in sheets_to_retry:
    if sheet_name not in wb.sheetnames:
        print(f"Advertencia: Hoja '{sheet_name}' no encontrada en Excel.")
        continue
        
    ws = wb[sheet_name]
    # Scan deeper: 50 rows
    rows = list(ws.iter_rows(values_only=True, max_row=50))
    if not rows: continue

    ruc_found = None
    distrito_found = None
    header_row_idx = -1
    col_map = {}

    # 1. Búsqueda agresiva de RUC y Header
    for i, row in enumerate(rows):
        row_str = " ".join([str(c) for c in row if c]).upper()
        norm_row = [normalize_text(c) for c in row]

        # RUC: Buscar en celdas individuales si no hallado
        if not ruc_found:
            for cell in row:
                if isinstance(cell, (int, str)):
                    c_str = str(cell).strip()
                    # Aceptar 10/20 seguido de 9 digitos
                    if re.match(r'^(10|20)\d{9}$', c_str):
                        ruc_found = c_str
                        break
            # Regex en fila completa
            if not ruc_found:
                 match = re.search(r'\b(10|20)\d{9}\b', row_str)
                 if match: ruc_found = match.group(0)

        # Distrito
        if not distrito_found:
            for dist in DISTRITOS:
                if dist in row_str:
                    distrito_found = dist
                    break
        
        # Header (Flexible)
        if header_row_idx == -1:
            is_header = False
            # Criterio 1: FECHA y (SERVICIO o PAGO)
            if 'FECHA' in row_str and ('SERVICIO' in row_str or 'PAGO' in row_str): is_header = True
            # Criterio 2: MES y SERVICIO
            if 'MES' in row_str and 'SERVICIO' in row_str: is_header = True
            # Criterio 3: F. y SERVICIO
            if 'F.' in row_str and 'SERVICIO' in row_str: is_header = True
            
            if is_header:
                 # Mapear
                for idx, val in enumerate(norm_row):
                    if ('FECHA' in val or 'F.' in val) and ('SERVICIO' in val or 'ATENCION' in val or 'VISITA' in val): col_map['FECHA_SERVICIO'] = idx
                    elif ('FECHA' in val or 'F.' in val) and 'PAGO' in val: col_map['FECHA_PAGO'] = idx
                    elif 'MES' in val: col_map['MES'] = idx
                    elif 'FORMA' in val: col_map['FORMA_PAGO'] = idx
                    elif 'DESC' in val or 'RESIDU' in val: col_map['DESCRIPCION'] = idx
                
                if 'FECHA_SERVICIO' in col_map:
                    header_row_idx = i

    # 2. Generar SQL si hay datos mínimos
    if valid_match := (ruc_found or sheet_name): # Siempre intentamos generar algo si hay header
        
        target_sede_var = f"@id_sede_retry_{count_recovered}"
        sql_statements.append(f"\n-- Hoja Retry: {sheet_name}")

        has_sede_logic = False
        if ruc_found:
             extra_cond = f"AND (s.distrito LIKE '%{distrito_found}%' OR s.direccion LIKE '%{distrito_found}%')" if distrito_found else ""
             sql_statements.append(f"""
SET {target_sede_var} = (SELECT s.id_sede FROM Sede s JOIN Empresa e ON s.id_empresa = e.id_empresa WHERE e.ruc = '{ruc_found}' {extra_cond} LIMIT 1);""")
             # Fallback RUC puro
             if extra_cond:
                 sql_statements.append(f"SET {target_sede_var} = COALESCE({target_sede_var}, (SELECT s.id_sede FROM Sede s JOIN Empresa e ON s.id_empresa = e.id_empresa WHERE e.ruc = '{ruc_found}' LIMIT 1));")
             has_sede_logic = True
        else:
             # Match por nombre de hoja (arriesgado pero pedido por usuario)
             clean_name = sheet_name.replace("'", "").strip()
             sql_statements.append(f"SET {target_sede_var} = (SELECT id_sede FROM Sede WHERE nombre_comercial LIKE '%{clean_name}%' LIMIT 1);")
             has_sede_logic = True

        # Procesar Servicios
        if header_row_idx != -1 and has_sede_logic:
            start_row = header_row_idx + 1
            services_in_sheet = 0
            for r_idx, row in enumerate(rows[start_row:], start=start_row+1):
                if not row: continue
                
                # Check bounds
                idx_fs = col_map.get('FECHA_SERVICIO')
                if idx_fs is None or idx_fs >= len(row): continue
                
                fecha_ejecucion = parse_date_excel(row[idx_fs])
                if not fecha_ejecucion: continue

                # Extraction logic (same as before)
                mes = clean_sql_val(row[col_map['MES']]) if 'MES' in col_map and col_map['MES'] < len(row) else 'NULL'
                
                fecha_pago = 'NULL'
                estado_pago = 'pendiente'
                if 'FECHA_PAGO' in col_map and col_map['FECHA_PAGO'] < len(row):
                    fp = parse_date_excel(row[col_map['FECHA_PAGO']])
                    if fp:
                        fecha_pago = f"'{fp}'"
                        estado_pago = 'pagado'
                
                forma_pago = clean_sql_val(row[col_map['FORMA_PAGO']]) if 'FORMA_PAGO' in col_map and col_map['FORMA_PAGO'] < len(row) else 'NULL'
                desc = normalize_residuo(row[col_map['DESCRIPCION']] if 'DESCRIPCION' in col_map and col_map['DESCRIPCION'] < len(row) else '')

                sql_statements.append(f"""
INSERT INTO Servicio (id_sede, id_planta, id_contrato, fecha_ejecucion, mes_servicio, fecha_pago, estado_pago, forma_pago, descripcion_residuo, estado)
SELECT {target_sede_var}, 1, (SELECT id_contrato FROM ContratoServicio WHERE id_sede = {target_sede_var} LIMIT 1), '{fecha_ejecucion}', {mes}, {fecha_pago}, '{estado_pago}', {forma_pago}, '{desc}', 'completado'
WHERE {target_sede_var} IS NOT NULL;""")
                services_in_sheet += 1
            
            if services_in_sheet > 0:
                count_recovered += 1
                count_services_generated += services_in_sheet
                recovered_sheets.append(sheet_name)
                print(f"Recuperada: {sheet_name} ({services_in_sheet} servicios)")
            else:
                msg = f"Header encontrado (L{header_row_idx+1}), pero filas de servicios invalidas (falta fecha o formato incorrecto)"
                print(f"Falló extracción servicios: {sheet_name} ({msg})")
                failed_reasons[sheet_name] = msg
        else:
             msg = f"No se encontró Header 'FECHA/SERVICIO' o RUC/Sede (RUC: {ruc_found}, HeaderIdx: {header_row_idx})"
             print(f"Falló header/sede: {sheet_name} ({msg})")
             failed_reasons[sheet_name] = msg


with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_statements))

# Write final failed list
with open('final_failed_sheets.txt', 'w', encoding='utf-8') as f:
    f.write(f"Total Inicial Fallidas: {len(sheets_to_retry)}\n")
    f.write(f"Recuperadas: {len(recovered_sheets)}\n")
    f.write(f"Total Final Fallidas: {len(sheets_to_retry) - len(recovered_sheets)}\n")
    f.write("-" * 30 + "\n")
    for s in sheets_to_retry:
        if s not in recovered_sheets:
            reason = failed_reasons.get(s, "Error desconocido")
            f.write(f"{s} | Motivo: {reason}\n")

print(f"\n✅ SQL Retry generado: {OUTPUT_SQL}")
print(f"📄 Reporte final de fallos: final_failed_sheets.txt")
print(f"Hojas recuperadas: {count_recovered}")
print(f"Servicios extra: {count_services_generated}")
