
import openpyxl
import re
from datetime import datetime

try:
    from unidecode import unidecode as unidecode_fn
except ImportError:
    def unidecode_fn(s):
        return s

# --- Configuración ---
INPUT_FILE = 'PAGOS CLIENTES LIMA.xlsx'
OUTPUT_SQL = 'control/database/migrations/import_servicios_2026.sql'
LOG_FILE = 'manual_review_needed.txt'

# --- Listas de ayuda ---
DISTRITOS = [
    'ANCON', 'ATE', 'BARRANCO', 'BREÑA', 'CARABAYLLO', 'CHACLACAYO', 'CHORRILLOS', 'CIENEGUILLA',
    'COMAS', 'EL AGUSTINO', 'INDEPENDENCIA', 'JESUS MARIA', 'LA MOLINA', 'LA VICTORIA', 'LIMA',
    'LINCE', 'LOS OLIVOS', 'LURIGANCHO', 'LURIN', 'MAGDALENA', 'MIRAFLORES', 'PACHACAMAC',
    'PUCUSANA', 'PUEBLO LIBRE', 'PUENTE PIEDRA', 'PUNTA HERMOSA', 'PUNTA NEGRA', 'RIMAC',
    'SAN BARTOLO', 'SAN BORJA', 'SAN ISIDRO', 'SAN JUAN DE LURIGANCHO', 'SAN JUAN DE MIRAFLORES',
    'SAN LUIS', 'SAN MARTIN DE PORRES', 'SAN MIGUEL', 'SANTA ANITA', 'SANTA MARIA DEL MAR',
    'SANTA ROSA', 'SANTIAGO DE SURCO', 'SURQUILLO', 'VILLA EL SALVADOR', 'VILLA MARIA DEL TRIUNFO'
]

# Formatos de fecha soportados (orden de prioridad)
DATE_FORMATS = [
    '%Y-%m-%d',
    '%d/%m/%Y',
    '%d-%m-%Y',
    '%d/%m/%y',   # año corto: 09/11/21
    '%d-%m-%y',   # año corto: 09-11-21
    '%m/%d/%Y',   # formato US fallback
    '%m/%d/%y',   # formato US fallback corto
]

def normalize_text(text):
    if not text: return ""
    return unidecode_fn(str(text)).upper().strip()

def clean_sql_val(val):
    if val is None: return 'NULL'
    val_str = str(val).strip().replace('\t', '').replace("'", "\\'")
    if not val_str: return 'NULL'
    return f"'{val_str}'"

def parse_date_excel(cell_val):
    """Parsear una celda que puede ser datetime o texto a formato YYYY-MM-DD."""
    if isinstance(cell_val, datetime):
        return cell_val.strftime('%Y-%m-%d')
    if not cell_val:
        return None
    
    # Limpiar whitespace y tabs
    val_str = str(cell_val).strip().replace('\t', '')
    if not val_str:
        return None
    
    # Tomar solo la parte de fecha (antes de espacio/hora)
    date_part = val_str.split(' ')[0]
    
    for fmt in DATE_FORMATS:
        try:
            dt = datetime.strptime(date_part, fmt)
            if 2000 <= dt.year <= 2100:
                return dt.strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None

def is_date_like(cell_val):
    """Verificar si un valor parece ser una fecha (sin parsear estrictamente)."""
    if isinstance(cell_val, datetime):
        return True
    if not cell_val:
        return False
    val_str = str(cell_val).strip()
    # Patrón básico: contiene números y separadores de fecha
    return bool(re.match(r'^\s*\d{1,4}[/\-\.]\d{1,2}[/\-\.]\d{1,4}\s*', val_str))

def normalize_residuo(desc):
    desc = normalize_text(desc)
    if 'BIO' in desc: return 'Biocontaminado'
    if 'ESPECIAL' in desc: return 'Especial'
    if 'CADAVER' in desc or 'ANIMAL' in desc: return 'Cadaver Animal'
    if 'MEDICAMENTO' in desc or 'VENCIDO' in desc: return 'Medicamento Vencido'
    return 'Otros'

def detect_date_column(data_rows, expected_col, total_cols):
    """
    Si la columna esperada no tiene fechas, buscar dinámicamente la columna correcta.
    Retorna el índice de la columna con más fechas válidas, o None.
    """
    # Primero verificar la columna esperada
    if expected_col is not None:
        dates_found = 0
        for row in data_rows[:10]:
            if row and expected_col < len(row):
                if parse_date_excel(row[expected_col]):
                    dates_found += 1
        if dates_found >= 1:
            return expected_col
    
    # Si la esperada falla, probar todas las columnas
    best_col = None
    best_count = 0
    for col_idx in range(min(total_cols, 15)):
        dates_found = 0
        for row in data_rows[:10]:
            if row and col_idx < len(row):
                if parse_date_excel(row[col_idx]):
                    dates_found += 1
        if dates_found > best_count:
            best_count = dates_found
            best_col = col_idx
    
    return best_col if best_count >= 1 else None

def find_header_row(rows):
    """
    Buscar la fila de encabezado con múltiples estrategias.
    Retorna (header_row_idx, col_map) o (-1, {}).
    """
    col_map = {}
    
    # Estrategia 1: Buscar "FECHA" + "SERVICIO" o "PAGO" en la misma fila
    for i, row in enumerate(rows[:15]):
        if not row:
            continue
        row_str = " ".join([str(c) for c in row if c]).upper()
        norm_row = [normalize_text(c) for c in row]
        
        if 'FECHA' in row_str and ('SERVICIO' in row_str or 'PAGO' in row_str):
            temp_map = {}
            for idx, val in enumerate(norm_row):
                if not val:
                    continue
                if 'FECHA' in val and ('SERVICIO' in val or 'EJECUCION' in val):
                    temp_map['FECHA_SERVICIO'] = idx
                elif 'FECHA' in val and 'PAGO' in val:
                    temp_map['FECHA_PAGO'] = idx
                elif val == 'MES' or 'MES DEL' in val or 'MES SERVICIO' in val:
                    temp_map['MES'] = idx
                elif 'FORMA' in val and 'PAGO' in val:
                    temp_map['FORMA_PAGO'] = idx
                elif 'DESC' in val or 'RESIDU' in val:
                    temp_map['DESCRIPCION'] = idx
                elif 'FACTURA' in val:
                    temp_map['FACTURA'] = idx
                elif 'MANIFIESTO' in val:
                    temp_map['MANIFIESTO'] = idx
                elif 'MONTO' in val or 'IMPORTE' in val or 'PRECIO' in val:
                    temp_map['MONTO'] = idx
            
            if 'FECHA_SERVICIO' in temp_map or 'FECHA_PAGO' in temp_map:
                col_map = temp_map
                return i, col_map
    
    # Estrategia 2: Buscar cualquier fila con "FECHA" y al menos 3 columnas no vacías
    for i, row in enumerate(rows[:15]):
        if not row:
            continue
        norm_row = [normalize_text(c) for c in row]
        non_empty = [v for v in norm_row if v.strip()]
        has_fecha = any('FECHA' in v for v in norm_row)
        
        if has_fecha and len(non_empty) >= 3:
            temp_map = {}
            for idx, val in enumerate(norm_row):
                if not val:
                    continue
                if 'FECHA' in val and ('INICIO' in val or not any(k in val for k in ['PAGO', 'FIN', 'VENCIMIENTO'])):
                    temp_map['FECHA_SERVICIO'] = idx
                elif 'FECHA' in val and 'PAGO' in val:
                    temp_map['FECHA_PAGO'] = idx
                elif 'MES' in val:
                    temp_map['MES'] = idx
            
            if temp_map:
                col_map = temp_map
                return i, col_map
    
    return -1, {}


# ===== MAIN =====
print(f"Cargando archivo '{INPUT_FILE}' (esto puede tardar)...")
wb = openpyxl.load_workbook(INPUT_FILE, data_only=True, read_only=True)

sql_statements = []
sql_statements.append("-- Importación de Servicios Históricos (PAGOS CLIENTES LIMA)")
sql_statements.append(f"-- Generado: {datetime.now()}")
sql_statements.append("SET FOREIGN_KEY_CHECKS = 0;")
sql_statements.append("-- Asegurar existencia de Planta Principal (Requerido por FK)")
sql_statements.append("INSERT IGNORE INTO Planta (id_planta, razon_social, ruc, direccion, activo) VALUES (1, 'PLANTA PRINCIPAL', '20000000000', 'LIMA', 1);")

log_lines = []

count_sheets_processed = 0
count_services_generated = 0
count_sheets_with_services = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True, max_row=500))  # Ampliado de 100 a 500
    
    if not rows: continue

    # 1. Identificar RUC y Posible Dirección en Encabezado (filas 0-14)
    ruc_found = None
    distrito_found = None
    
    for i, row in enumerate(rows[:15]):
        row_str = " ".join([str(c) for c in row if c]).upper()
        
        # Buscar RUC
        if not ruc_found:
            match = re.search(r'\b(10|20)\d{9}\b', row_str)
            if match:
                ruc_found = match.group(0)
            else:
                for cell in row:
                    if isinstance(cell, (int, str)) and re.match(r'^(10|20)\d{9}$', str(cell).strip()):
                        ruc_found = str(cell).strip()
                        break
        
        # Buscar Distrito
        if not distrito_found:
            row_norm = normalize_text(row_str)
            for dist in DISTRITOS:
                if dist in row_norm:
                    distrito_found = dist
                    break

    # 2. Buscar Header
    header_row_idx, col_map = find_header_row(rows)

    # 3. Generar lógica de selección de ID Sede
    sql_statements.append(f"\n-- Hoja: {sheet_name}")
    
    target_sede_var = f"@id_sede_{count_sheets_processed}"
    
    if ruc_found:
        if distrito_found:
            sql_statements.append(f"""
SET {target_sede_var} = (
    SELECT s.id_sede FROM Sede s 
    JOIN Empresa e ON s.id_empresa = e.id_empresa 
    WHERE e.ruc = '{ruc_found}' 
    AND (s.distrito LIKE '%{distrito_found}%' OR s.direccion LIKE '%{distrito_found}%')
    LIMIT 1
);""")
            sql_statements.append(f"""
SET {target_sede_var} = COALESCE({target_sede_var}, (SELECT s.id_sede FROM Sede s JOIN Empresa e ON s.id_empresa = e.id_empresa WHERE e.ruc = '{ruc_found}' LIMIT 1));""")
        else:
            sql_statements.append(f"SET {target_sede_var} = (SELECT s.id_sede FROM Sede s JOIN Empresa e ON s.id_empresa = e.id_empresa WHERE e.ruc = '{ruc_found}' LIMIT 1);")
    else:
        clean_name = sheet_name.replace("'", "").strip()
        sql_statements.append(f"SET {target_sede_var} = (SELECT id_sede FROM Sede WHERE nombre_comercial LIKE '%{clean_name}%' LIMIT 1);")
        log_lines.append(f"REVISION REQUERIDA: Hoja '{sheet_name}' sin RUC. Se intentó match por nombre.")

    # 4. Procesar Servicios
    if header_row_idx != -1:
        start_row = header_row_idx + 1
        data_rows = rows[start_row:]
        
        if not data_rows:
            log_lines.append(f"INFO: Hoja '{sheet_name}' tiene encabezados pero no hay filas de datos.")
            count_sheets_processed += 1
            continue
        
        # Detección inteligente de columna de fecha
        expected_fecha_col = col_map.get('FECHA_SERVICIO')
        total_cols = max(len(r) for r in data_rows if r) if data_rows else 0
        
        actual_fecha_col = detect_date_column(data_rows, expected_fecha_col, total_cols)
        
        # Si no se encontró fecha en la columna esperada, intentar con FECHA_PAGO como alternativa
        if actual_fecha_col is None and 'FECHA_PAGO' in col_map:
            actual_fecha_col = detect_date_column(data_rows, col_map['FECHA_PAGO'], total_cols)
            if actual_fecha_col is not None:
                log_lines.append(f"INFO: Hoja '{sheet_name}' - usando FECHA_PAGO (col {actual_fecha_col}) como fecha de ejecución.")
        
        if actual_fecha_col is None:
            log_lines.append(f"WARN: Hoja '{sheet_name}' tiene header (fila {header_row_idx+1}) pero ninguna columna contiene fechas parseables. col_map={col_map}")
            count_sheets_processed += 1
            continue
        
        # Si la columna real difiere de la esperada, recalibrar col_map desplazando
        if expected_fecha_col is not None and actual_fecha_col != expected_fecha_col:
            offset = actual_fecha_col - expected_fecha_col
            log_lines.append(f"INFO: Hoja '{sheet_name}' - datos desplazados por {offset} columnas. Recalibrando.")
            new_map = {}
            for key, idx in col_map.items():
                new_idx = idx + offset
                if 0 <= new_idx < total_cols:
                    new_map[key] = new_idx
            col_map = new_map
            col_map['FECHA_SERVICIO'] = actual_fecha_col
        elif expected_fecha_col is None:
            col_map['FECHA_SERVICIO'] = actual_fecha_col

        curr_services = 0
        for r_idx, row in enumerate(data_rows):
            if not row: continue
            
            # Obtener Fecha Servicio
            idx_fs = col_map.get('FECHA_SERVICIO')
            if idx_fs is None or idx_fs >= len(row): continue
            
            fecha_ejecucion = parse_date_excel(row[idx_fs])
            if not fecha_ejecucion: continue  # Skip filas sin fecha válida
            
            # Mes
            mes = 'NULL'
            if 'MES' in col_map and col_map['MES'] < len(row):
                mes = clean_sql_val(row[col_map['MES']])
            
            # Fecha pago y estado
            fecha_pago = 'NULL'
            estado_pago = 'pendiente'
            if 'FECHA_PAGO' in col_map and col_map['FECHA_PAGO'] < len(row):
                fp = parse_date_excel(row[col_map['FECHA_PAGO']])
                if fp:
                    fecha_pago = f"'{fp}'"
                    estado_pago = 'pagado'
            
            # Forma de pago
            forma_pago = 'NULL'
            if 'FORMA_PAGO' in col_map and col_map['FORMA_PAGO'] < len(row):
                forma_pago = clean_sql_val(row[col_map['FORMA_PAGO']])
            
            # Descripción residuo
            desc_raw = ''
            if 'DESCRIPCION' in col_map and col_map['DESCRIPCION'] < len(row):
                desc_raw = row[col_map['DESCRIPCION']]
            descripcion = normalize_residuo(desc_raw)

            # Factura y Manifiesto (limpiar tabs y .0)
            factura = None
            if 'FACTURA' in col_map and col_map['FACTURA'] < len(row):
                val = row[col_map['FACTURA']]
                if val:
                    cleaned = str(val).strip().replace('\t', '').replace('.0', '')
                    if cleaned:
                        factura = cleaned
            
            manifiesto = None
            if 'MANIFIESTO' in col_map and col_map['MANIFIESTO'] < len(row):
                val = row[col_map['MANIFIESTO']]
                if val:
                    cleaned = str(val).strip().replace('\t', '').replace('.0', '')
                    if cleaned:
                        manifiesto = cleaned
            
            # Generar INSERT
            sql_statements.append(f"""
INSERT INTO Servicio (id_sede, id_planta, id_contrato, fecha_ejecucion, mes_servicio, fecha_pago, estado_pago, forma_pago, descripcion_residuo, estado)
SELECT {target_sede_var}, 1, (SELECT id_contrato FROM ContratoServicio WHERE id_sede = {target_sede_var} LIMIT 1), '{fecha_ejecucion}', {mes}, {fecha_pago}, '{estado_pago}', {forma_pago}, '{descripcion}', 'completado'
WHERE {target_sede_var} IS NOT NULL;""")
            
            # Insertar Factura vinculada
            if factura:
                factura_escaped = factura.replace("'", "\\'")
                sql_statements.append(f"""
INSERT INTO Factura (id_servicio, numero_factura) 
SELECT LAST_INSERT_ID(), '{factura_escaped}' 
WHERE {target_sede_var} IS NOT NULL AND LAST_INSERT_ID() > 0;""")

            # Insertar Manifiesto vinculado
            if manifiesto:
                manifiesto_escaped = manifiesto.replace("'", "\\'")
                tipo_res = descripcion if descripcion else 'General'
                sql_statements.append(f"""
INSERT INTO Manifiesto (id_servicio, numero_manifiesto, tipo_residuo, peso_kg) 
SELECT LAST_INSERT_ID(), '{manifiesto_escaped}', '{tipo_res}', 0 
WHERE {target_sede_var} IS NOT NULL AND LAST_INSERT_ID() > 0;""")

            curr_services += 1
        
        count_services_generated += curr_services
        if curr_services > 0:
            count_sheets_with_services += 1
        else:
            log_lines.append(f"INFO: Hoja '{sheet_name}' tiene encabezados pero no se extrajeron servicios (fecha_col={actual_fecha_col}, col_map={col_map}).")
    else:
        log_lines.append(f"ERROR: Hoja '{sheet_name}' no se detectó encabezado de servicios.")

    count_sheets_processed += 1
    if count_sheets_processed % 100 == 0:
        print(f"Procesadas {count_sheets_processed} hojas... ({count_services_generated} servicios)")

sql_statements.append("SET FOREIGN_KEY_CHECKS = 1;")

# Escribir Outputs
with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql_statements))

with open(LOG_FILE, 'w', encoding='utf-8') as f:
    f.write('\n'.join(log_lines))

print(f"\n{'='*60}")
print(f"✅ SQL generado en: {OUTPUT_SQL}")
print(f"📄 Log de revisión: {LOG_FILE}")
print(f"📊 Hojas procesadas: {count_sheets_processed}")
print(f"📊 Hojas con servicios: {count_sheets_with_services}")
print(f"📊 Total Servicios Generados: {count_services_generated}")
print(f"⚠️  Entradas en log: {len(log_lines)}")
