import openpyxl
import re
from datetime import datetime

try:
    from unidecode import unidecode as unidecode_fn
except ImportError:
    def unidecode_fn(s):
        return s

# --- Configuración ---
INPUT_FILE = 'PAGOS CLIENTES LIMA.xlsx'
OUTPUT_SQL = 'control/database/migrations/import_docs_only_2026.sql'
LOG_FILE = 'manual_review_needed_docs.txt'

DISTRITOS = [
    'ANCON', 'ATE', 'BARRANCO', 'BREÑA', 'CARABAYLLO', 'CHACLACAYO', 'CHORRILLOS', 'CIENEGUILLA',
    'COMAS', 'EL AGUSTINO', 'INDEPENDENCIA', 'JESUS MARIA', 'LA MOLINA', 'LA VICTORIA', 'LIMA',
    'LINCE', 'LOS OLIVOS', 'LURIGANCHO', 'LURIN', 'MAGDALENA', 'MIRAFLORES', 'PACHACAMAC',
    'PUCUSANA', 'PUEBLO LIBRE', 'PUENTE PIEDRA', 'PUNTA HERMOSA', 'PUNTA NEGRA', 'RIMAC',
    'SAN BARTOLO', 'SAN BORJA', 'SAN ISIDRO', 'SAN JUAN DE LURIGANCHO', 'SAN JUAN DE MIRAFLORES',
    'SAN LUIS', 'SAN MARTIN DE PORRES', 'SAN MIGUEL', 'SANTA ANITA', 'SANTA MARIA DEL MAR',
    'SANTA ROSA', 'SANTIAGO DE SURCO', 'SURQUILLO', 'VILLA EL SALVADOR', 'VILLA MARIA DEL TRIUNFO'
]

DATE_FORMATS = [
    '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y', '%m/%d/%Y', '%m/%d/%y'
]

DATE_MIN_YEAR = 2019
DATE_MAX_YEAR = 2026

def normalize_text(text):
    if not text: return ""
    return unidecode_fn(str(text)).upper().strip()

def clean_sql_val(val):
    if val is None: return 'NULL'
    val_str = str(val).strip().replace('\t', '').replace("'", "\\'")
    if not val_str: return 'NULL'
    return f"'{val_str}'"

def parse_date_excel(cell_val):
    if isinstance(cell_val, datetime):
        if DATE_MIN_YEAR <= cell_val.year <= DATE_MAX_YEAR:
            return cell_val.strftime('%Y-%m-%d')
        return None
    if not cell_val: return None
    val_str = str(cell_val).strip().replace('\t', '')
    if not val_str: return None
    date_part = val_str.split(' ')[0]
    for fmt in DATE_FORMATS:
        try:
            dt = datetime.strptime(date_part, fmt)
            if DATE_MIN_YEAR <= dt.year <= DATE_MAX_YEAR:
                return dt.strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None

def normalize_residuo(desc):
    desc = normalize_text(desc)
    if 'BIO' in desc: return 'Biocontaminado'
    if 'ESPECIAL' in desc: return 'Especial'
    if 'CADAVER' in desc or 'ANIMAL' in desc: return 'Cadaver Animal'
    if 'MEDICAMENTO' in desc or 'VENCIDO' in desc: return 'Medicamento Vencido'
    return 'Otros'

def detect_date_column(data_rows, expected_col, total_cols):
    if expected_col is not None:
        dates_found = sum(1 for row in data_rows[:10] if row and expected_col < len(row) and parse_date_excel(row[expected_col]))
        if dates_found >= 1: return expected_col
    best_col, best_count = None, 0
    for col_idx in range(min(total_cols, 15)):
        dates_found = sum(1 for row in data_rows[:10] if row and col_idx < len(row) and parse_date_excel(row[col_idx]))
        if dates_found > best_count:
            best_count = dates_found
            best_col = col_idx
    return best_col if best_count >= 1 else None

def find_header_row(rows):
    col_map = {}
    for i, row in enumerate(rows[:15]):
        if not row: continue
        row_str = " ".join([str(c) for c in row if c]).upper()
        norm_row = [normalize_text(c) for c in row]
        if 'FECHA' in row_str and ('SERVICIO' in row_str or 'PAGO' in row_str):
            for idx, val in enumerate(norm_row):
                if not val: continue
                if 'FECHA' in val and ('SERVICIO' in val or 'EJECUCION' in val): col_map['FECHA_SERVICIO'] = idx
                elif 'FECHA' in val and 'PAGO' in val: col_map['FECHA_PAGO'] = idx
                elif val == 'MES' or 'MES DEL' in val or 'MES SERVICIO' in val: col_map['MES'] = idx
                elif 'DESC' in val or 'RESIDU' in val: col_map['DESCRIPCION'] = idx
                elif 'FACTURA' in val: col_map['FACTURA'] = idx
                elif 'MANIFIESTO' in val: col_map['MANIFIESTO'] = idx
                elif 'GUIA' in val: col_map['GUIA'] = idx
            if 'FECHA_SERVICIO' in col_map or 'FECHA_PAGO' in col_map: return i, col_map
    for i, row in enumerate(rows[:15]):
        if not row: continue
        norm_row = [normalize_text(c) for c in row]
        if any('FECHA' in v for v in norm_row) and len([v for v in norm_row if v.strip()]) >= 3:
            for idx, val in enumerate(norm_row):
                if not val: continue
                if 'FECHA' in val and ('INICIO' in val or not any(k in val for k in ['PAGO', 'FIN', 'VENCIMIENTO'])): col_map['FECHA_SERVICIO'] = idx
                elif 'FECHA' in val and 'PAGO' in val: col_map['FECHA_PAGO'] = idx
                elif 'FACTURA' in val: col_map['FACTURA'] = idx
                elif 'MANIFIESTO' in val: col_map['MANIFIESTO'] = idx
                elif 'GUIA' in val: col_map['GUIA'] = idx
            if col_map: return i, col_map
    return -1, {}

print(f"Cargando archivo '{INPUT_FILE}'...")
wb = openpyxl.load_workbook(INPUT_FILE, data_only=True, read_only=True)

sql = [
    "-- Importación de Facturas, Manifiestos y Guías ÚNICAMENTE",
    f"-- Generado: {datetime.now()}",
    "SET FOREIGN_KEY_CHECKS = 0;"
]
log_lines = []

count_sheets = 0
count_facturas = 0
count_manifiestos = 0
count_guias = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True, max_row=500))
    if not rows: continue

    ruc_found, distrito_found = None, None
    for row in rows[:15]:
        row_str = " ".join([str(c) for c in row if c]).upper()
        if not ruc_found:
            match = re.search(r'\b(10|20)\d{9}\b', row_str)
            ruc_found = match.group(0) if match else next((str(c).strip() for c in row if isinstance(c, (int, str)) and re.match(r'^(10|20)\d{9}$', str(c).strip())), None)
        if not distrito_found:
            distrito_found = next((dist for dist in DISTRITOS if dist in normalize_text(row_str)), None)

    header_row_idx, col_map = find_header_row(rows)
    if header_row_idx == -1:
        log_lines.append(f"ERROR: No headers in {sheet_name}")
        continue

    target_sede_var = f"@id_sede_{count_sheets}"
    sql.append(f"\n-- Hoja: {sheet_name}")
    
    if ruc_found:
        if distrito_found:
            sql.append(f"SET {target_sede_var} = (SELECT s.id_sede FROM Sede s JOIN Empresa e ON s.id_empresa = e.id_empresa WHERE e.ruc = '{ruc_found}' AND (s.distrito LIKE '%{distrito_found}%' OR s.direccion LIKE '%{distrito_found}%') LIMIT 1);")
            sql.append(f"SET {target_sede_var} = COALESCE({target_sede_var}, (SELECT s.id_sede FROM Sede s JOIN Empresa e ON s.id_empresa = e.id_empresa WHERE e.ruc = '{ruc_found}' LIMIT 1));")
        else:
            sql.append(f"SET {target_sede_var} = (SELECT s.id_sede FROM Sede s JOIN Empresa e ON s.id_empresa = e.id_empresa WHERE e.ruc = '{ruc_found}' LIMIT 1);")
    else:
        clean_name = sheet_name.replace("'", "").strip()
        sql.append(f"SET {target_sede_var} = (SELECT id_sede FROM Sede WHERE nombre_comercial LIKE '%{clean_name}%' LIMIT 1);")

    data_rows = rows[header_row_idx + 1:]
    actual_fecha_col = detect_date_column(data_rows, col_map.get('FECHA_SERVICIO'), max(len(r) for r in data_rows if r) if data_rows else 0)
    if actual_fecha_col is None and 'FECHA_PAGO' in col_map:
        actual_fecha_col = detect_date_column(data_rows, col_map['FECHA_PAGO'], max(len(r) for r in data_rows if r))

    if actual_fecha_col is None:
        log_lines.append(f"WARN: No dates in {sheet_name}")
        count_sheets += 1
        continue

    for row in data_rows:
        if not row or actual_fecha_col >= len(row): continue
        fecha = parse_date_excel(row[actual_fecha_col])
        if not fecha: continue

        factura, manifiesto, guia = None, None, None
        
        if 'FACTURA' in col_map and col_map['FACTURA'] < len(row) and row[col_map['FACTURA']]:
            f = str(row[col_map['FACTURA']]).strip().replace('\t', '').replace('.0', '')
            if f and f.lower() not in ['pendiente', 'no', 'sin', 'sn', '-', 'na', 'n/a', 'x']: factura = f
            
        if 'MANIFIESTO' in col_map and col_map['MANIFIESTO'] < len(row) and row[col_map['MANIFIESTO']]:
            m = str(row[col_map['MANIFIESTO']]).strip().replace('\t', '').replace('.0', '')
            if m and m.lower() not in ['pendiente', 'no', 'sin', 'sn', '-', 'na', 'n/a', 'x']: manifiesto = m

        if 'GUIA' in col_map and col_map['GUIA'] < len(row) and row[col_map['GUIA']]:
            g = str(row[col_map['GUIA']]).strip().replace('\t', '').replace('.0', '')
            if g and g.lower() not in ['pendiente', 'no', 'sin', 'sn', '-', 'na', 'n/a', 'x']: guia = g

        has_any = factura or manifiesto or guia
        if not has_any:
            continue

        desc = normalize_residuo(row[col_map['DESCRIPCION']] if 'DESCRIPCION' in col_map and col_map['DESCRIPCION'] < len(row) else '')

        # To link documents to the service, find id_servicio first, then insert.
        # Fallback to mes_servicio if exact date is off in the Excel.
        y_m = fecha[:7] # YYYY-MM
        
        sql.append(f"""
SET @id_svc_{count_sheets} = (SELECT id_servicio FROM Servicio WHERE id_sede = {target_sede_var} AND (fecha_ejecucion = '{fecha}' OR mes_servicio LIKE '%{y_m}%' OR mes_servicio LIKE '%{int(fecha[5:7])}%') LIMIT 1);""")
        
        if factura:
            sql.append(f"INSERT INTO Factura (id_servicio, numero_factura) SELECT @id_svc_{count_sheets}, '{factura}' WHERE @id_svc_{count_sheets} IS NOT NULL;")
            count_facturas += 1
            
        if manifiesto:
            sql.append(f"INSERT INTO Manifiesto (id_servicio, numero_manifiesto, tipo_residuo, peso_kg) SELECT @id_svc_{count_sheets}, '{manifiesto}', '{desc}', 0 WHERE @id_svc_{count_sheets} IS NOT NULL;")
            count_manifiestos += 1
            
        if guia:
            sql.append(f"INSERT INTO Guia (id_servicio, numero_guia) SELECT @id_svc_{count_sheets}, '{guia}' WHERE @id_svc_{count_sheets} IS NOT NULL;")
            count_guias += 1

    count_sheets += 1
    if count_sheets % 50 == 0:
        print(f"Hojas: {count_sheets} | Docs generados: Facturas={count_facturas}, Manifiestos={count_manifiestos}, Guías={count_guias}")

sql.append("SET FOREIGN_KEY_CHECKS = 1;")

with open(OUTPUT_SQL, 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql))

with open(LOG_FILE, 'w', encoding='utf-8') as f:
    f.write('\n'.join(log_lines))

print(f"\n✅ SQL generado en: {OUTPUT_SQL}")
print(f"📄 Docs generados: {count_facturas} Facturas, {count_manifiestos} Manifiestos, {count_guias} Guias")
